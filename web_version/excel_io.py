import pandas as pd, os, csv, zipfile
import xml.etree.ElementTree as ET
import openpyxl

def fast_xlsx_sheets(file_path):
    """Extremely fast sheet name extractor for .xlsx using zipfile and XML parsing."""
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            with z.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                # Namespaces can vary, but sheet names are in 'sheet' tags under 'sheets'
                sheets = []
                for sheet in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                    name = sheet.get('name')
                    if name: sheets.append(name)
                return sheets
    except:
        return []

def _sniff_csv(file_path, enc):
    # ... (existing code)
    try:
        with open(file_path,'r',encoding=enc) as f:
            sample=f.read(4096)
            dialect=csv.Sniffer().sniff(sample, delimiters=[',','\t','|',';'])
            return dialect.delimiter
    except:
        return None

# ... (rest of file)

import shutil
import tempfile
import uuid

class SafeExcelReader:
    """
    Context manager to handle locked Excel files on Windows.
    If the file is locked, it copies it to a temp file and reads that.
    """
    def __init__(self, file_path):
        self.original_path = file_path
        self.temp_path = None
        self.use_temp = False

    def __enter__(self):
        if not os.path.exists(self.original_path):
            return self.original_path

        # Try to open normally first?
        # Actually, just try to check if it's readable.
        try:
            with open(self.original_path, 'rb') as f:
                pass
            return self.original_path
        except PermissionError:
            # File is locked, copy to temp
            self.use_temp = True
            ext = os.path.splitext(self.original_path)[1]
            self.temp_path = os.path.join(tempfile.gettempdir(), f"em_temp_{uuid.uuid4()}{ext}")
            try:
                shutil.copy2(self.original_path, self.temp_path)
                return self.temp_path
            except Exception as e:
                print(f"Failed to copy locked file: {e}")
                return self.original_path # Fallback to original, might fail again

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.use_temp and self.temp_path and os.path.exists(self.temp_path):
            try:
                os.remove(self.temp_path)
            except: pass

def get_sheet_names(file_path):
    if not file_path or not os.path.exists(file_path):
        return []
    
    with SafeExcelReader(file_path) as path_to_read:
        try:
            ext = os.path.splitext(path_to_read)[1].lower()
            if ext == '.xlsx':
                names = fast_xlsx_sheets(path_to_read)
                if names: return names
                try:
                    wb = openpyxl.load_workbook(path_to_read, read_only=True, keep_links=False)
                    names = wb.sheetnames
                    wb.close()
                    return names
                except:
                    return pd.ExcelFile(path_to_read).sheet_names
            elif ext == '.xls':
                return pd.ExcelFile(path_to_read).sheet_names
            elif ext == '.csv':
                return ['CSV']
        except Exception as e:
            print(f"Sheet load error ({file_path}): {e}")
            return []
    return []

def read_header_file(file_path, sheet_name=0, header_row=1):
    # header_row is 1-based, we need 0-based for pandas
    try:
        with SafeExcelReader(file_path) as path_to_read:
            ext = os.path.splitext(path_to_read)[1].lower()
            header_idx = header_row - 1
            
            if ext == '.xlsx':
                # Use openpyxl read_only for speed
                try:
                    wb = openpyxl.load_workbook(path_to_read, read_only=True, data_only=True)
                    # Handle sheet index or name
                    target_sheet = sheet_name
                    if isinstance(target_sheet, int):
                        target_sheet = wb.sheetnames[target_sheet]
                    ws = wb[target_sheet]
                    
                    # Iterate rows to find header
                    for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
                        if i == 0: # First row from iterator is our header
                            headers = [str(c).strip() if c is not None else f"Unnamed: {idx}" for idx, c in enumerate(row)]
                            wb.close()
                            return headers
                    wb.close()
                    return []
                except Exception as e:
                    print(f"Fast header read failed, falling back: {e}")
                    # Fallback to pandas
                    pass
    
            if ext in ['.xls','.xlsx']:
                df=pd.read_excel(path_to_read, sheet_name=sheet_name, header=header_idx, nrows=0)
            elif ext=='.csv':
                df=None
                for enc in ['cp949','utf-8','euc-kr']:
                    try:
                        sep=_sniff_csv(path_to_read, enc)
                        df=pd.read_csv(path_to_read, header=header_idx, nrows=0, encoding=enc, sep=sep, engine='python')
                        break
                    except: 
                        continue
                if df is None: return []
            else:
                return []
            return [str(c).strip() for c in df.columns.tolist()]
    except:
        return []

def read_table_file(file_path, sheet_name, header_row, usecols):
    with SafeExcelReader(file_path) as path_to_read:
        ext=os.path.splitext(path_to_read)[1].lower()
        header_idx=header_row-1
        if isinstance(usecols,str): usecols=[usecols]
        usecols=[str(c).strip() for c in (usecols or [])]
        if ext == '.xlsx':
            try:
                # Breakthrough: Calamine is significantly faster for large XLSX
                df = pd.read_excel(path_to_read, sheet_name=sheet_name, header=header_idx, engine='calamine')
            except:
                df = pd.read_excel(path_to_read, sheet_name=sheet_name, header=header_idx)
        elif ext == '.xls':
            df = pd.read_excel(path_to_read, sheet_name=sheet_name, header=header_idx)
        elif ext == '.csv':
            df = None
            for enc in ['utf-8-sig', 'cp949', 'utf-8', 'euc-kr']:
                try:
                    sep = _sniff_csv(path_to_read, enc)
                    # Expert: Engine 'c' is faster than 'python'
                    df = pd.read_csv(path_to_read, header=header_idx, encoding=enc, sep=sep, engine='c', low_memory=False)
                    break
                except: continue
            if df is None:
                raise Exception("CSV 파일 인코딩/구분자를 인식하지 못했습니다.")
        else:
            return pd.DataFrame()
        df.columns=[str(c).strip() for c in df.columns]
        
        if usecols:
            existing=[c for c in usecols if c in df.columns]
            missing=[c for c in usecols if c not in df.columns]
            df=df[existing] if existing else pd.DataFrame(index=df.index)
            for c in missing: df[c]=""
            df=df.reindex(columns=usecols, fill_value="")
        
        df=df.astype(str).replace(['nan','NaN','None','<NA>'],'')
        return df

def get_unique_values(file_path, sheet_name, header_row, column_name, progress_callback=None):
    """
    Returns a sorted list of unique entries for a specific column.
    Optimized for .xlsx using iter_rows (read_only).
    """
    try:
        with SafeExcelReader(file_path) as path_to_read:
            ext = os.path.splitext(path_to_read)[1].lower()
            
            if ext == '.xlsx':
                try:
                    wb = openpyxl.load_workbook(path_to_read, read_only=True, data_only=True)
                    target_sheet = sheet_name
                    if isinstance(target_sheet, int):
                        target_sheet = wb.sheetnames[target_sheet]
                    ws = wb[target_sheet]
    
                    # Find column index
                    col_idx = -1
                    # Read header row
                    for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
                        for idx, val in enumerate(row):
                            if str(val).strip() == column_name:
                                col_idx = idx
                                break
                        break
                    
                    if col_idx == -1: 
                        wb.close()
                        return []
    
                    unique_set = set()
                    excluded = {'nan', 'none', 'null', ''}
                    
                    # OPTIMIZED LOOP: Minimize object creation and string calls for 1M rows
                    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), 1):
                        # Report progress every 5000 rows
                        if progress_callback and row_idx % 5000 == 0:
                            progress_callback(row_idx)
                            
                        if col_idx < len(row_vals):
                            val = row_vals[col_idx]
                            if val is not None:
                                if isinstance(val, str):
                                    s_val = val.strip()
                                    if s_val and s_val.lower() not in excluded:
                                        unique_set.add(s_val)
                                else:
                                    s_val = str(val).strip()
                                    if s_val and s_val.lower() not in excluded:
                                        unique_set.add(s_val)
                    
                    wb.close()
                    unique_list = sorted(list(unique_set))
                    # Report final count
                    if progress_callback: progress_callback(row_idx)
                    return ["(값 있음)", "(값 없음)"] + unique_list
                except Exception as e:
                    print(f"Fast unique read failed, falling back: {e}")
    
            # Fallback to pandas (slow but reliable)
            header_idx = header_row - 1
            if ext in ['.xls', '.xlsx']:
                df = pd.read_excel(path_to_read, sheet_name=sheet_name, header=header_idx, usecols=[column_name])
            elif ext == '.csv':
                df = None
                for enc in ['cp949', 'utf-8', 'euc-kr']:
                    try:
                        sep = _sniff_csv(path_to_read, enc)
                        df = pd.read_csv(path_to_read, header=header_idx, usecols=[column_name], encoding=enc, sep=sep, engine='python')
                        break
                    except: continue
            else:
                return []
                
            if df is not None and not df.empty:
                series = df[column_name].astype(str).str.strip()
                unique_vals = sorted([v for v in series.unique() if v and v.lower() not in ['nan', 'none', 'null']])
                # Prepend Expert Options
                return ["(값 있음)", "(값 없음)"] + unique_vals
                
            return ["(값 있음)", "(값 없음)"]
    except Exception as e:
        print(f"Unique value load error ({column_name}): {e}")
        return []

def write_xlsx(file_path, df, sheet_name="Sheet1"):
    try:
        try:
            writer=pd.ExcelWriter(file_path, engine='xlsxwriter')
        except:
            writer=pd.ExcelWriter(file_path, engine='openpyxl')
        with writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except PermissionError:
        raise Exception(f"파일 저장 실패: 권한 부족 또는 파일이 열려있습니다.\n'{os.path.basename(file_path)}' 파일을 닫고 다시 시도하세요.")
    except Exception as e:
        raise Exception(f"파일 저장 실패: {e}")

