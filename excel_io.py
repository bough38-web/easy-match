import pandas as pd, os, csv
import openpyxl

def _sniff_csv(file_path, enc):
    # ... (existing code)
    try:
        with open(file_path,'r',encoding=enc) as f:
            sample=f.read(4096)
            dialect=csv.Sniffer().sniff(sample, delimiters=[',','\t','|',';'])
            return dialect.delimiter
    except:
        return None

# ... (rest of file)

def get_sheet_names(file_path):
    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.xlsx':
            try:
                return openpyxl.load_workbook(file_path, read_only=True, keep_links=False).sheetnames
            except:
                return pd.ExcelFile(file_path).sheet_names
        elif ext == '.xls':
            return pd.ExcelFile(file_path).sheet_names
        elif ext == '.csv':
            return ['CSV']
    except Exception as e:
        print(f"Sheet load error: {e}")
        return []

def read_header_file(file_path, sheet_name=0, header_row=1):
    try:
        ext = os.path.splitext(file_path)[1].lower()
        header_idx = header_row - 1
        
        if ext == '.xlsx':
            # Use openpyxl read_only for speed
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                # Handle sheet index or name
                if isinstance(sheet_name, int):
                    sheet_name = wb.sheetnames[sheet_name]
                ws = wb[sheet_name]
                
                # Iterate rows to find header
                for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
                    if i == 0: # First row from iterator is our header
                        headers = [str(c).strip() if c is not None else f"Unnamed: {idx}" for idx, c in enumerate(row)]
                        wb.close()
                        return headers
                wb.close()
                return []
            except Exception as e:
                print(f"Fast header read failed, falling back: {e}")
                # Fallback to pandas
                pass

        if ext in ['.xls','.xlsx']:
            df=pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx, nrows=0)
        elif ext=='.csv':
            df=None
            for enc in ['cp949','utf-8','euc-kr']:
                try:
                    sep=_sniff_csv(file_path, enc)
                    df=pd.read_csv(file_path, header=header_idx, nrows=0, encoding=enc, sep=sep, engine='python')
                    break
                except: 
                    continue
            if df is None: return []
        else:
            return []
        return [str(c).strip() for c in df.columns.tolist()]
    except:
        return []

def read_table_file(file_path, sheet_name, header_row, usecols):
    ext=os.path.splitext(file_path)[1].lower()
    header_idx=header_row-1
    if isinstance(usecols,str): usecols=[usecols]
    usecols=[str(c).strip() for c in (usecols or [])]
    if ext in ['.xls','.xlsx']:
        df=pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)
    elif ext=='.csv':
        df=None
        for enc in ['cp949','utf-8','euc-kr']:
            try:
                sep=_sniff_csv(file_path, enc)
                df=pd.read_csv(file_path, header=header_idx, encoding=enc, sep=sep, engine='python')
                break
            except: 
                continue
        if df is None:
            raise Exception("CSV 파일 인코딩/구분자를 인식하지 못했습니다.")
    else:
        return pd.DataFrame()
    df.columns=[str(c).strip() for c in df.columns]
    existing=[c for c in usecols if c in df.columns]
    missing=[c for c in usecols if c not in df.columns]
    df=df[existing] if existing else pd.DataFrame()
    for c in missing: df[c]=""
    df=df.reindex(columns=usecols, fill_value="")
    df=df.astype(str).replace(['nan','NaN','None','<NA>'],'')
    return df

def get_unique_values(file_path, sheet_name, header_row, column_name):
    """
    Returns a sorted list of unique entries for a specific column.
    Optimized for .xlsx using iter_rows (read_only).
    """
    try:
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.xlsx':
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if isinstance(sheet_name, int):
                    sheet_name = wb.sheetnames[sheet_name]
                ws = wb[sheet_name]

                # Find column index
                col_idx = -1
                # Read header row
                for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
                    for idx, val in enumerate(row):
                        if str(val).strip() == column_name:
                            col_idx = idx + 1 # 1-based index for iter_cols? No, iter_rows returns tuple. 
                            # actually iter_cols is not fully supported in read_only mode efficiently?
                            # It says "Warning: iter_cols is slow in read_only".
                            # Better to iterate rows and pick the index.
                            col_idx = idx
                            break
                    break
                
                if col_idx == -1: 
                    wb.close()
                    return []

                unique_set = set()
                # Pre-fetch lowercase check set for speed
                excluded = {'nan', 'none', 'null', ''}
                
                # OPTIMIZED LOOP: Minimize object creation and string calls for 1M rows
                for row_vals in ws.iter_rows(min_row=header_row+1, values_only=True):
                    if col_idx < len(row_vals):
                        val = row_vals[col_idx]
                        if val is not None:
                            # Faster check: skip if it's already a string and likely not empty
                            if isinstance(val, str):
                                s_val = val.strip()
                                if s_val and s_val.lower() not in excluded:
                                    unique_set.add(s_val)
                            else:
                                s_val = str(val).strip()
                                if s_val and s_val.lower() not in excluded:
                                    unique_set.add(s_val)
                
                wb.close()
                unique_list = sorted(list(unique_set))
                return ["(값 있음)", "(값 없음)"] + unique_list
            except Exception as e:
                print(f"Fast unique read failed, falling back: {e}")

        # Fallback to pandas (slow but reliable)
        header_idx = header_row - 1
        if ext in ['.xls', '.xlsx']:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx, usecols=[column_name])
        elif ext == '.csv':
            df = None
            for enc in ['cp949', 'utf-8', 'euc-kr']:
                try:
                    sep = _sniff_csv(file_path, enc)
                    df = pd.read_csv(file_path, header=header_idx, usecols=[column_name], encoding=enc, sep=sep, engine='python')
                    break
                except: continue
        else:
            return []
            
        if df is not None and not df.empty:
            series = df[column_name].astype(str).str.strip()
            unique_vals = sorted([v for v in series.unique() if v and v.lower() not in ['nan', 'none', 'null']])
            # Prepend Expert Options
            return ["(값 있음)", "(값 없음)"] + unique_vals
            
        return ["(값 있음)", "(값 없음)"]
    except Exception as e:
        print(f"Unique value load error ({column_name}): {e}")
        return []

def write_xlsx(file_path, df, sheet_name="Sheet1"):
    try:
        try:
            writer=pd.ExcelWriter(file_path, engine='xlsxwriter')
        except:
            writer=pd.ExcelWriter(file_path, engine='openpyxl')
        with writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except PermissionError:
        raise Exception(f"파일 저장 실패: 권한 부족 또는 파일이 열려있습니다.\n'{os.path.basename(file_path)}' 파일을 닫고 다시 시도하세요.")
    except Exception as e:
        raise Exception(f"파일 저장 실패: {e}")

