from __future__ import annotations

import os
import json
import sys
import tkinter as tk
import traceback
from tkinter import ttk, filedialog, messagebox, simpledialog
from PIL import Image, ImageTk

# Try to import tkinterdnd2 for drag and drop support
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DRAG_DROP_AVAILABLE = True
except ImportError:
    DRAG_DROP_AVAILABLE = False

from __version__ import __version__
from diagnostics import collect_summary, format_summary
from excel_io import read_header_file, get_sheet_names
from open_excel import (
    list_open_books,
    list_sheets,
    read_header_open,
    xlwings_available,
)
from matcher import match_universal
from commercial_config import (
    ADMIN_PASSWORD,
    CREATOR_NAME,
    SALES_INFO,
    BANK_INFO,
    PRICE_INFO,
    CONTACT_INFO,
    TRIAL_EMAIL,
    TRIAL_SUBJECT,
)
from admin_panel import AdminPanel

# -------------------------
# Matched Data Preview Dialog
# -------------------------
def show_preview_dialog(parent, out_path, summary, preview_df):
    dialog = tk.Toplevel(parent)
    dialog.title("작업 완료 및 데이터 미리보기")
    dialog.geometry("900x650")
    dialog.transient(parent)
    dialog.grab_set()

    # Glassmorphism-style header in dialog
    header = GradientFrame(dialog, color1="#2c3e50", color2="#1a2a3a", height=60)
    header.pack(fill="x")
    header.create_text(20, 30, text="매칭 결과 미리보기 (상위 5행)", font=(get_system_font()[0], 18, "bold"), fill="white", anchor="w")

    main_frame = ttk.Frame(dialog, padding=20)
    main_frame.pack(fill="both", expand=True)

    # Summary box with accent border
    summary_frame = tk.Frame(main_frame, bg="#f8f9fa", highlightbackground="#dee2e6", highlightthickness=1)
    summary_frame.pack(fill="x", pady=(0, 20))
    
    inner_summary = tk.Frame(summary_frame, bg="#f8f9fa", padx=15, pady=15)
    inner_summary.pack(fill="x")
    
    ttk.Label(inner_summary, text=summary, font=(get_system_font()[0], 12, "bold"), background="#f8f9fa").pack(anchor="w")
    ttk.Label(inner_summary, text=f"저장 위치: {os.path.basename(out_path)}", font=(get_system_font()[0], 10), foreground="#2B579A", background="#f8f9fa").pack(anchor="w", pady=(5, 0))

    # Table
    table_frame = ttk.Frame(main_frame)
    table_frame.pack(fill="both", expand=True)

    cols = list(preview_df.columns)
    tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
    
    # Add scrolls
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    for col in cols:
        tree.heading(col, text=col)
        # Approximate width based on content
        tree.column(col, width=120, minwidth=80)

    for _, row in preview_df.iterrows():
        tree.insert("", "end", values=[str(v) for v in row])

    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    table_frame.grid_columnconfigure(0, weight=1)
    table_frame.grid_rowconfigure(0, weight=1)

    # Buttons
    btn_frame = ttk.Frame(main_frame)
    btn_frame.pack(fill="x", pady=(20, 0))
    
    def open_folder():
        folder = os.path.dirname(out_path)
        if sys.platform == "darwin":
            import subprocess
            subprocess.run(["open", folder])
        else:
            os.startfile(folder)

    ttk.Button(btn_frame, text="폴더 열기", command=open_folder).pack(side="left", padx=5)
    ttk.Button(btn_frame, text="확인 (닫기)", command=dialog.destroy).pack(side="right", padx=5)

    dialog.wait_window()

APP_TITLE = "Easy Match(이지 매치)"
APP_DESCRIPTION = "엑셀과 CSV를 하나로, 클릭 한 번으로 끝나는 데이터 매칭"

# 출력 디렉토리를 사용자의 Documents 폴더로 설정 (macOS 호환성)
if sys.platform == "darwin":  # macOS
    OUT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "EasyMatch_Outputs")
else:  # Windows
    OUT_DIR = os.path.join(os.getcwd(), "outputs")

from config import PRESET_FILE, REPLACE_FILE, get_system_font
os.makedirs(OUT_DIR, exist_ok=True)


# -------------------------
# Gradient Frame (Glassmorphism Effect)
# Using a sleek, semi-transparent feeling gradient
# -------------------------
class GradientFrame(tk.Canvas):
    def __init__(self, master, color1="#1a2a3a", color2="#2c3e50", **kwargs):
        super().__init__(master, highlightthickness=0, **kwargs)
        self.color1 = color1
        self.color2 = color2
        self.bind("<Configure>", self._draw_gradient)

    def _draw_gradient(self, event=None):
        self.delete("gradient")
        width = self.winfo_width()
        height = self.winfo_height()
        if width == 0 or height == 0: return

        (r1,g1,b1) = self.winfo_rgb(self.color1)
        (r2,g2,b2) = self.winfo_rgb(self.color2)
        
        # 0-65535 -> 0-255
        r1, g1, b1 = r1>>8, g1>>8, b1>>8
        r2, g2, b2 = r2>>8, g2>>8, b2>>8

        for i in range(height):
            # Interpolate
            r = int(r1 + (r2 - r1) * i / height)
            g = int(g1 + (g2 - g1) * i / height)
            b = int(b1 + (b2 - b1) * i / height)
            color = "#%02x%02x%02x" % (r, g, b)
            self.create_line(0, i, width, i, tags=("gradient",), fill=color)
        
        # Add Glass Highlight (Top)
        self.create_line(0, 0, width, 0, fill="#5d6d7e", width=1, tags=("gradient",))
        # Add Shadow (Bottom)
        self.create_line(0, height-1, width, height-1, fill="#1c2833", width=1, tags=("gradient",))
        
        self.tag_lower("gradient")


# -------------------------
# ToolTip Class
# -------------------------
class ToolTip:
    """
    Create a tooltip for a given widget with improved styling and positioning.
    """
    def __init__(self, widget, text, delay=500):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.tooltip_window = None
        self.id_after = None
        
        widget.bind("<Enter>", self.on_enter)
        widget.bind("<Leave>", self.on_leave)
        widget.bind("<Button>", self.on_leave)  # Hide on click
    
    def on_enter(self, event=None):
        # Schedule tooltip display after delay
        self.id_after = self.widget.after(self.delay, self.show_tooltip)
    
    def on_leave(self, event=None):
        # Cancel scheduled tooltip
        if self.id_after:
            self.widget.after_cancel(self.id_after)
            self.id_after = None
        # Hide tooltip if visible
        self.hide_tooltip()
    
    def show_tooltip(self):
        if self.tooltip_window:
            return
        
        # Get widget position
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        
        # Create tooltip window
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        
        # Create tooltip label with modern styling
        label = tk.Label(
            self.tooltip_window,
            text=self.text,
            background="#2c3e50",
            foreground="#ecf0f1",
            relief="solid",
            borderwidth=1,
            font=(get_system_font()[0], 10),
            padx=8,
            pady=5,
            justify="left"
        )
        label.pack()
    
    def hide_tooltip(self):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


# -------------------------
# Replace rules persistence
# -------------------------
def _load_replace_file():
    try:
        if os.path.exists(REPLACE_FILE):
            with open(REPLACE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)

            # 구버전 호환: {"presetName": {...}} 형태였던 경우
            if isinstance(data, dict) and "presets" not in data:
                return {"presets": data, "active": {}, "active_name": ""}

            return {
                "presets": data.get("presets", {}) or {},
                "active": data.get("active", {}) or {},
                "active_name": data.get("active_name", "") or "",
            }
    except Exception:
        pass

    return {"presets": {}, "active": {}, "active_name": ""}


def _save_replace_file(presets: dict, active: dict, active_name: str = ""):
    with open(REPLACE_FILE, "w", encoding="utf-8") as f:
        json.dump(
            {
                "presets": presets or {},
                "active": active or {},
                "active_name": active_name or "",
            },
            f,
            ensure_ascii=False,
            indent=4,
        )


class ReplacementEditor(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("데이터 치환 규칙 설정")
        self.geometry("650x520")

        data = _load_replace_file()
        self.presets = data["presets"]
        self.rules = data["active"] or {}
        self.active_name = data.get("active_name", "") or ""

        top_frame = ttk.LabelFrame(self, text="규칙 프리셋 관리", padding=10)
        top_frame.pack(fill="x", padx=10, pady=5)

        self.preset_var = tk.StringVar()
        self.cb_preset = ttk.Combobox(
            top_frame, textvariable=self.preset_var, state="readonly", width=25
        )
        self.cb_preset.pack(side="left", padx=5)
        self.cb_preset.bind("<<ComboboxSelected>>", self.on_preset_selected)

        ttk.Button(top_frame, text="저장 (Save)", command=self.save_preset).pack(
            side="left", padx=2
        )
        ttk.Button(top_frame, text="삭제 (Del)", command=self.delete_preset).pack(
            side="left", padx=2
        )
        ttk.Button(top_frame, text="초기화 (Reset)", command=self.clear_all).pack(
            side="right", padx=2
        )

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)

        form_frame = ttk.Frame(main_frame)
        form_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(form_frame, text="대상 컬럼명:").grid(
            row=0, column=0, padx=5, sticky="w"
        )
        self.ent_col = ttk.Entry(form_frame, width=20)
        self.ent_col.grid(row=0, column=1, padx=5)

        ttk.Label(form_frame, text="변경 전 (Old):").grid(
            row=0, column=2, padx=5, sticky="w"
        )
        self.ent_old = ttk.Entry(form_frame, width=15)
        self.ent_old.grid(row=0, column=3, padx=5)

        ttk.Label(form_frame, text="변경 후 (New):").grid(
            row=0, column=4, padx=5, sticky="w"
        )
        self.ent_new = ttk.Entry(form_frame, width=15)
        self.ent_new.grid(row=0, column=5, padx=5)

        ttk.Button(form_frame, text="추가/수정", command=self.add_rule).grid(
            row=0, column=6, padx=10
        )

        columns = ("col", "old", "new")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings", height=15)
        self.tree.heading("col", text="대상 컬럼")
        self.tree.heading("old", text="변경 전 값 (Old)")
        self.tree.heading("new", text="변경 후 값 (New)")
        self.tree.column("col", width=180)
        self.tree.column("old", width=180)
        self.tree.column("new", width=180)

        scroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self.on_double_click)

        bot_frame = ttk.Frame(self, padding=10)
        bot_frame.pack(fill="x")
        ttk.Label(
            bot_frame, text="* 목록을 더블클릭하면 삭제됩니다.", foreground="gray"
        ).pack(side="left")
        ttk.Button(bot_frame, text="닫기 (적용)", command=self._close).pack(side="right")

        self.update_preset_cb()
        self.refresh_tree()

    def _persist(self):
        name = self.cb_preset.get() if self.cb_preset.get() in self.presets else self.active_name
        _save_replace_file(self.presets, self.rules, name)

    def _close(self):
        self._persist()
        self.destroy()

    def add_rule(self):
        col = self.ent_col.get().strip()
        old = self.ent_old.get().strip()
        new = self.ent_new.get().strip()
        if not col or not old:
            messagebox.showwarning("입력 오류", "컬럼명과 변경 전 값은 필수입니다.")
            return
        self.rules.setdefault(col, {})[old] = new
        self.refresh_tree()
        self.ent_old.delete(0, tk.END)
        self.ent_new.delete(0, tk.END)
        self.ent_old.focus()
        self._persist()

    def refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for col, mapping in (self.rules or {}).items():
            if not isinstance(mapping, dict):
                continue
            for old, new in mapping.items():
                self.tree.insert("", "end", values=(col, old, new))

    def on_double_click(self, event=None):
        item = self.tree.selection()
        if not item:
            return
        col, old, new = self.tree.item(item, "values")
        if messagebox.askyesno("삭제", f"'{col}' 컬럼의 '{old}' -> '{new}' 규칙을 삭제할까요?"):
            if col in self.rules and old in self.rules[col]:
                del self.rules[col][old]
                if not self.rules[col]:
                    del self.rules[col]
            self.refresh_tree()
            self._persist()

    def clear_all(self):
        if messagebox.askyesno("초기화", "모든 규칙을 지우시겠습니까?"):
            self.rules = {}
            self.refresh_tree()
            self._persist()

    def update_preset_cb(self):
        names = list((self.presets or {}).keys())
        self.cb_preset["values"] = names
        self.cb_preset.set("설정을 선택하세요" if names else "저장된 설정 없음")

    def save_preset(self):
        if not self.rules:
            messagebox.showwarning("경고", "저장할 규칙이 없습니다.")
            return
        name = simpledialog.askstring("설정 저장", "치환 규칙 세트 이름 입력:")
        if not name:
            return
        name = name.strip()
        if not name:
            return
        import copy

        self.presets[name] = copy.deepcopy(self.rules)
        self.active_name = name
        self.cb_preset.set(name)
        self.update_preset_cb()
        self._persist()
        messagebox.showinfo("저장", f"[{name}] 규칙이 저장되었습니다.")

    def delete_preset(self):
        name = self.cb_preset.get()
        if name in self.presets and messagebox.askyesno(
            "삭제", f"[{name}] 규칙 세트를 삭제하시겠습니까?"
        ):
            del self.presets[name]
            if self.active_name == name:
                self.active_name = ""
            self.update_preset_cb()
            self._persist()

    def on_preset_selected(self, event=None):
        name = self.cb_preset.get()
        if name in self.presets:
            import copy

            self.rules = copy.deepcopy(self.presets[name])
            self.active_name = name
            self.refresh_tree()
            self._persist()

    def get_rules(self):
        return self.rules


# -------------------------
# Grid Check List (Moved from below to allow inheritance)
# -------------------------
class GridCheckList(ttk.Frame):
    def __init__(self, master, columns=4, height=300):
        super().__init__(master)
        self.columns = columns
        self.vars: dict[str, tk.BooleanVar] = {}

        top = ttk.Frame(self)
        top.pack(fill="x", pady=(0, 4))
        self.q = tk.StringVar()
        ent = ttk.Entry(top, textvariable=self.q)
        ent.pack(side="left", fill="x", expand=True)
        ent.bind("<KeyRelease>", self._on_key_release)
        self.search_timer = None
        ttk.Button(top, text="지우기", width=6, command=self._clear).pack(side="left", padx=4)

        self.canvas = tk.Canvas(self, height=height, bg="white", highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.configure(yscrollcommand=sb.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfigure(self.inner_id, width=e.width))

        self.all_items: list[str] = []

    def set_items(self, items):
        items = list(items or [])
        if self.all_items == items: 
            return # No change, skip expensive re-render
            
        self.all_items = items
        self.vars.clear()
        for it in self.all_items:
            self.vars[it] = tk.BooleanVar(value=False)
        self._render(self.all_items)

    def _render(self, items):
        # Clear current widgets
        for w in self.inner.winfo_children():
            w.destroy()
            
        # LIMIT RENDERING for performance
        MAX_ITEMS = 300
        display_items = items[:MAX_ITEMS]
        
        if not display_items:
            ttk.Label(self.inner, text="(데이터 없음)").pack()
            return
            
        # Fast batch creation: withdraw window updates during child creation
        self.inner.pack_forget() 
        
        for idx, it in enumerate(display_items):
            if it not in self.vars:
                self.vars[it] = tk.BooleanVar(value=False)
                
            # Use tk.Checkbutton for significantly better PERFORMANCE than ttk in large lists
            cb = tk.Checkbutton(self.inner, text=it, variable=self.vars[it], 
                                bg="white", activebackground="white", selectcolor="white",
                                font=(get_system_font()[0], 10))
            cb.grid(row=idx // self.columns, column=idx % self.columns, sticky="w", padx=4, pady=1)
            
        self.inner.pack(fill="both", expand=True)
            
        # Show warning if truncated
        if len(items) > MAX_ITEMS:
            remain = len(items) - MAX_ITEMS
            msg = f"...외 {remain}개 항목 (검색하여 찾으세요)"
            lbl = ttk.Label(self.inner, text=msg, foreground="gray", font=(get_system_font()[0], 9))
            lbl.grid(row=(len(display_items) // self.columns) + 1, column=0, columnspan=self.columns, sticky="w", padx=5, pady=5)

        for i in range(self.columns):
            self.inner.columnconfigure(i, weight=1)

    def checked(self):
        return [k for k, v in self.vars.items() if v.get()]

    def check_all(self):
        for v in self.vars.values():
            v.set(True)

    def uncheck_all(self):
        for v in self.vars.values():
            v.set(False)

    def set_checked_items(self, items):
        self.uncheck_all()
        cnt = 0
        for it in items or []:
            if it in self.vars:
                self.vars[it].set(True)
                cnt += 1
        return cnt

    def _clear(self):
        self.q.set("")
        self._render(self.all_items)

    def _on_key_release(self, event=None):
        if self.search_timer:
            self.after_cancel(self.search_timer)
        self.search_timer = self.after(300, self._filter)

    def _filter(self):
        self.search_timer = None
        q = (self.q.get() or "").strip().lower()
        if not q:
            self._render(self.all_items)
            return
            
        filtered = [it for it in self.all_items if q in it.lower()]
        self._render(filtered)


class MultiSelectListBox(GridCheckList):
    """Wrapper around GridCheckList to replace the old vertical Listbox"""
    def __init__(self, master, columns=4, height=150):
        # Allow passing height in pixels
        super().__init__(master, columns=columns, height=height)

    def get_selected(self):
        # Maintain compatibility with old Listbox API
        return self.checked()  

    def select_item(self, item):
        # Maintain compatibility with old Listbox API
        self.set_checked_items([item])






# Column Selector Dialog
# -------------------------

# Determine base class for App
if DRAG_DROP_AVAILABLE:
    BaseApp = TkinterDnD.Tk
else:
    BaseApp = tk.Tk

# -------------------------
# New UI Components for Redesign
# -------------------------
class MultiFilterRow:
    def __init__(self, master, get_cols_func, fetch_vals_func, on_remove):
        self.frame = ttk.Frame(master)
        self.frame.pack(fill="x", pady=2)
        
        self.get_cols = get_cols_func
        self.fetch_vals = fetch_vals_func
        
        self.col_var = tk.StringVar()
        self.op_var = tk.StringVar(value="==")
        self.val_var = tk.StringVar()
        
        # Column
        self.cb_col = ttk.Combobox(self.frame, textvariable=self.col_var, state="readonly", width=12)
        self.cb_col.pack(side="left", padx=(0, 5))
        self.cb_col.bind("<<ComboboxSelected>>", self._on_col_change)
        
        # Operator
        self.cb_op = ttk.Combobox(self.frame, textvariable=self.op_var, values=["==", ">=", "<=", ">", "<"], state="readonly", width=4)
        self.cb_op.pack(side="left", padx=(0, 5))
        
        # Value
        self.cb_val = ttk.Combobox(self.frame, textvariable=self.val_var, state="normal", width=12)
        self.cb_val.pack(side="left", padx=(0, 5))
        self.cb_val.set("(값 선택)")
        
        # Remove
        btn_rem = ttk.Button(self.frame, text="X", width=2, command=lambda: on_remove(self))
        btn_rem.pack(side="left")

        # Load Values Button (Lazy Load)
        self.btn_load = ttk.Button(self.frame, text="▼", width=2, command=self._load_values_async)
        self.btn_load.pack(side="left", padx=(2, 0))

        self.refresh_cols()

    def _on_op_change(self, event=None):
        # Kept for potential logic, though currently unbound
        pass

    def refresh_cols(self):
        cols = self.get_cols()
        self.cb_col["values"] = cols
        if cols and not self.col_var.get():
            self.cb_col.current(0)
            # Remove auto-fetch to prevent freeze
            # self._on_col_change() 

    def _on_col_change(self, event=None):
        # Clear values on column change, but do NOT auto-fetch
        self.cb_val.set("")
        self.cb_val["values"] = []

    def _load_values_async(self):
        col = self.col_var.get()
        if not col or not self.fetch_vals: return
        
        self.cb_val.set("Loading...")
        self.btn_load.state(["disabled"])
        
        import threading
        def _task():
            try:
                vals = self.fetch_vals(col)
                # Schedule UI update on main thread
                self.frame.after(0, lambda: self._update_vals(vals))
            except Exception as e:
                print(f"Error fetching values: {e}")
                self.frame.after(0, lambda: self._update_vals([], error=True))

        t = threading.Thread(target=_task, daemon=True)
        t.start()

    def _update_vals(self, vals, error=False):
        try:
            if not self.frame.winfo_exists(): return
            self.btn_load.state(["!disabled"])
            
            if error:
                self.cb_val.set("(오류 발생)")
                self.cb_val["values"] = []
            elif not vals:
                self.cb_val.set("(값 없음)")
                self.cb_val["values"] = []
            else:
                self.cb_val["values"] = vals
                if len(vals) > 0:
                    self.cb_val.set(vals[0])
                    # Remove flaky event_generate which might confuse the UI focus
                    # Instead, just set the value. User can click if they want to see more.
                    self.cb_val.current(0) 
        except Exception as e:
            print(f"UI Update Error: {e}")

    def get_config(self):
        return {
            "col": self.col_var.get(),
            "op": self.op_var.get(),
            "keyword": self.val_var.get()
        }

class FileLoaderFrame(ttk.LabelFrame):
    def __init__(self, master, title, mode_var, on_change=None, on_fetch_vals=None):
        super().__init__(master, text=title, padding=10)
        self.mode = mode_var
        self.on_change = on_change
        self.on_fetch_vals = on_fetch_vals
        
        self.path = tk.StringVar()
        self.book = tk.StringVar()
        self.sheet = tk.StringVar()
        self.header = tk.IntVar(value=1)

        # Top: Mode selection
        top = ttk.Frame(self)
        top.pack(fill="x", pady=(0, 5))
        
        ttk.Radiobutton(top, text="파일 불러오기", value="file", variable=self.mode, command=self._refresh_ui).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(top, text="열려있는 엑셀", value="open", variable=self.mode, command=self._refresh_ui).pack(side="left")
        
        # Align Refresh button
        ttk.Button(top, text="새로고침", command=self.refresh_all, width=8).pack(side="right")

        # File Mode UI
        self.f_frame = ttk.Frame(self)
        entry = ttk.Entry(self.f_frame, textvariable=self.path)
        entry.pack(side="left", fill="x", expand=True)
        # Drag Drop bind
        if DRAG_DROP_AVAILABLE:
            try:
                entry.drop_target_register(DND_FILES)
                entry.dnd_bind('<<Drop>>', self._on_drop)
            except: pass

        # Align Find button
        ttk.Button(self.f_frame, text="찾기", command=self._pick_file, width=8).pack(side="right", padx=(5, 0))

        # Open Mode UI
        self.o_frame = ttk.Frame(self)
        self.cb_book = ttk.Combobox(self.o_frame, textvariable=self.book, state="readonly")
        self.cb_book.pack(side="left", fill="x", expand=True)
        self.cb_book.bind("<<ComboboxSelected>>", self._on_book_select)

        # Common UI (Sheet/Header)
        self.c_frame = ttk.Frame(self)
        self.c_frame.pack(fill="x", pady=(5, 0))
        
        ttk.Label(self.c_frame, text="시트:").pack(side="left")
        self.cb_sheet = ttk.Combobox(self.c_frame, textvariable=self.sheet, state="readonly", width=15)
        self.cb_sheet.pack(side="left", padx=5)
        self.cb_sheet.bind("<<ComboboxSelected>>", self._notify_change)
        
        ttk.Label(self.c_frame, text="헤더:").pack(side="left", padx=(10, 0))
        ttk.Spinbox(self.c_frame, from_=1, to=99, textvariable=self.header, width=5, command=self._notify_change).pack(side="left", padx=5)

        # Filter UI (Redesigned)
        self.f_opt_frame = ttk.Frame(self)
        self.f_opt_frame.pack(fill="x", pady=(5, 0))
        
        self.filter_expanded = tk.BooleanVar(value=False)
        self.btn_toggle_f = ttk.Button(self.f_opt_frame, text="▶ 필터 조건 설정 (0개)", command=self._toggle_filters)
        self.btn_toggle_f.pack(side="left")
        
        self.f_container = ttk.Frame(self)
        # Initially hidden
        
        self.f_rows: List[MultiFilterRow] = []
        
        self.row_controls = ttk.Frame(self.f_container)
        self.row_controls.pack(fill="x", pady=5)
        
        self.btn_add_f = ttk.Button(self.row_controls, text="+ 필터 추가", command=self._add_filter_row)
        self.btn_add_f.pack(side="left")
        
        self._refresh_ui()

    def _toggle_filters(self):
        if self.filter_expanded.get():
            self.f_container.pack_forget()
            self.filter_expanded.set(False)
            self._update_filter_btn_text()
        else:
            self.f_container.pack(fill="x", pady=5, after=self.f_opt_frame)
            self.filter_expanded.set(True)
            self._update_filter_btn_text()

    def _update_filter_btn_text(self):
        symbol = "▼" if self.filter_expanded.get() else "▶"
        count = len(self.f_rows)
        self.btn_toggle_f.config(text=f"{symbol} 필터 조건 설정 ({count}개)")

    def _add_filter_row(self):
        if len(self.f_rows) >= 5:
            messagebox.showwarning("제한", "필터 조건은 최대 5개까지 가능합니다.")
            return
        
        row = MultiFilterRow(self.f_container, self.get_cols_callback, self.on_fetch_vals, self._remove_filter_row)
        # Insert before controls
        row.frame.pack(after=None, before=self.row_controls)
        self.f_rows.append(row)
        self._update_filter_btn_text()

    def _remove_filter_row(self, row):
        row.frame.destroy()
        if row in self.f_rows:
            self.f_rows.remove(row)
        self._update_filter_btn_text()

    def get_cols_callback(self):
        # Helper for FilterRow to get current columns using App's cache
        cfg = self.get_config()
        app = self.winfo_toplevel()
        try:
            if cfg["type"] == "file" and cfg["path"]:
                if hasattr(app, "_fetch_headers"):
                    return app._fetch_headers(cfg["path"], cfg["sheet"], cfg["header"])
                from excel_io import read_header_file
                return read_header_file(cfg["path"], cfg["sheet"], cfg["header"])
            elif cfg["type"] == "open" and cfg["book"]:
                return read_header_open(cfg["book"], cfg["sheet"], cfg["header"])
        except: pass
        return []

    def _refresh_filter_cols(self):
        for row in self.f_rows:
            row.refresh_cols()

    def get_filters(self):
        return [r.get_config() for r in self.f_rows]

    def _notify_change(self, event=None):
        self._refresh_filter_cols()
        # Clear main app's unique value cache if configuration changes
        main_app = self.winfo_toplevel()
        if hasattr(main_app, "_clear_unique_cache"):
            main_app._clear_unique_cache()
            
        if self.on_change: self.on_change()

    def set_filter_state(self, active, col="", kw="", op="=="):
        """Programmatically set one filter (Legacy support / preset support)"""
        if active:
            if not self.filter_expanded.get():
                self._toggle_filters()
            # Clear existing if any? Usually presets want to set NEW state
            self.clear_filters()
            
            row = MultiFilterRow(self.f_container, self.get_cols_callback, self.on_fetch_vals, self._remove_filter_row)
            row.frame.pack(after=None, before=self.row_controls)
            self.f_rows.append(row)
            row.col_var.set(col)
            row.op_var.set(op)
            row.val_var.set(kw)
            self._update_filter_btn_text()

    def clear_filters(self):
        """Removes all filter rows and updates UI"""
        for r in self.f_rows[:]:
            r.frame.destroy()
        self.f_rows.clear()
        self._update_filter_btn_text()
        if self.filter_expanded.get():
            self._toggle_filters() # Collapse if expanded

    def refresh_all(self):
        """Refresh open books and clear filters"""
        self.refresh_open()
        self.clear_filters()
        if self.on_change: self.on_change()

    def _on_drop(self, event):
        try:
            file_path = event.data.strip('{}')
            valid_extensions = ('.xlsx', '.xls', '.csv')
            if not file_path.lower().endswith(valid_extensions):
                messagebox.showwarning("파일 형식 오류", f"지원하지 않는 파일 형식입니다.")
                return
            self.path.set(file_path)
            try:
                # Use master's cache if available
                app = self.winfo_toplevel()
                sheets = []
                if hasattr(app, "_fetch_sheet_names"):
                    sheets = app._fetch_sheet_names(file_path)
                else:
                    sheets = get_sheet_names(file_path)
                
                self.cb_sheet["values"] = sheets
                if sheets: self.cb_sheet.current(0)
            except Exception as e:
                messagebox.showerror("오류", f"시트 목록을 불러올 수 없습니다:\n{e}")
            
            event.widget.config(background="white")
            self._notify_change()
        except Exception as e:
            messagebox.showerror("오류", f"파일을 불러올 수 없습니다:\n{str(e)}")

    def _refresh_ui(self):
        if self.mode.get() == "file":
            self.o_frame.pack_forget()
            self.f_frame.pack(fill="x", before=self.c_frame)
        else:
            self.f_frame.pack_forget()
            self.o_frame.pack(fill="x", before=self.c_frame)
            self.refresh_open()
        self._notify_change()

    def _pick_file(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls *.csv"), ("All", "*.*")])
        if not p: return
        self.path.set(p)
        try:
            app = self.winfo_toplevel()
            sheets = []
            if hasattr(app, "_fetch_sheet_names"):
                sheets = app._fetch_sheet_names(p)
            else:
                sheets = get_sheet_names(p)
            
            self.cb_sheet["values"] = sheets
            if sheets: self.cb_sheet.current(0)
        except Exception as e:
            messagebox.showerror("오류", f"시트 목록을 불러올 수 없습니다:\n{e}")
        self._notify_change()

    def refresh_open(self):
        if self.mode.get() != "open": return
        if not xlwings_available():
            messagebox.showwarning("오류", "xlwings가 필요합니다.")
            return
        books = list_open_books()
        self.cb_book["values"] = books
        if books:
            if not self.book.get() or self.book.get() not in books:
                self.book.set(books[0])
        else:
            self.book.set("")
            self.cb_sheet["values"] = []
        self._on_book_select()

    def _on_book_select(self, event=None):
        if not self.book.get(): return
        try:
            sheets = list_sheets(self.book.get())
            self.cb_sheet["values"] = sheets
            if sheets: self.cb_sheet.current(0)
        except: pass
        self._notify_change()

    def _notify_change(self, event=None):
        self._refresh_filter_cols()
        # Clear main app's unique value cache if configuration changes
        main_app = self.winfo_toplevel()
        if hasattr(main_app, "_clear_unique_cache"):
            main_app._clear_unique_cache()
            
        if self.on_change: self.on_change()

    def get_config(self):
        return {
            "type": self.mode.get(),
            "path": self.path.get(),
            "book": self.book.get(),
            "sheet": self.sheet.get(),
            "header": int(self.header.get()),
        }


class ColumnSelectorFrame(ttk.LabelFrame):
    def __init__(self, master, title, height=150):
        super().__init__(master, text=title, padding=5)
        
        # Tools (Select All/None)
        tools = ttk.Frame(self)
        tools.pack(fill="x", pady=(0, 5))
        
        ttk.Button(tools, text="[V] 전체 선택", command=self.check_all, width=12).pack(side="left", padx=(0, 2))
        ttk.Button(tools, text="[X] 선택 해제", command=self.uncheck_all, width=12).pack(side="left", padx=(2, 0))
        
        # List
        self.list = GridCheckList(self, columns=3, height=height)
        self.list.pack(fill="both", expand=True)
        
    def check_all(self):
        self.list.check_all()
        
    def uncheck_all(self):
        self.list.uncheck_all()
        
    def set_items(self, items):
        self.list.set_items(items)
        
    def get_selected(self):
        return self.list.checked()


class TargetAdvFilterRow(ttk.Frame):
    def __init__(self, master, available_cols, on_remove, on_fetch_vals):
        super().__init__(master)
        self.on_fetch_vals = on_fetch_vals
        
        self.col_var = tk.StringVar()
        self.val_var = tk.StringVar()
        
        # Column selection
        self.cb_col = ttk.Combobox(self, textvariable=self.col_var, values=available_cols, state="readonly", width=15)
        self.cb_col.pack(side="left", padx=2)
        self.cb_col.bind("<<ComboboxSelected>>", self._on_col_change)
        
        # Value selection (dropdown)
        self.cb_val = ttk.Combobox(self, textvariable=self.val_var, state="readonly", width=15)
        self.cb_val.pack(side="left", padx=2)
        self.cb_val.set("(값 선택)")
        
        # Remove button
        btn_rem = ttk.Button(self, text="-", width=3, command=on_remove)
        btn_rem.pack(side="left", padx=2)

        # Load Values Button (Lazy Load)
        self.btn_load = ttk.Button(self, text="▼", width=2, command=self._load_values_async)
        self.btn_load.pack(side="left", padx=(2, 0))

    def _on_col_change(self, event=None):
        # Reset value when column changes
        self.cb_val.set("(값 선택)")
        self.cb_val["values"] = []

    def _load_values_async(self):
        col = self.col_var.get()
        if not col or not self.on_fetch_vals: return
        
        self.cb_val.set("Loading...")
        self.btn_load.state(["disabled"])
        
        import threading
        def _task():
            try:
                vals = self.on_fetch_vals(col)
                def _update():
                    if not self.winfo_exists(): return
                    self.btn_load.state(["!disabled"])
                    self.cb_val["values"] = vals
                    if vals: self.cb_val.current(0)
                    else: self.cb_val.set("(데이터 없음)")
                self.after(0, _update)
            except Exception as e:
                print(f"Error fetching values: {e}")
                self.after(0, lambda: self.btn_load.state(["!disabled"]))

        t = threading.Thread(target=_task, daemon=True)
        t.start()

    def get_filter(self):
        col = self.col_var.get()
        val = self.val_var.get()
        if col and val and val != "(값 선택)" and val != "(데이터 없음)":
            return {"col": col, "values": [val]}
        return None


class TargetFilterFrame(ttk.LabelFrame):
    def __init__(self, master, get_cols_func, get_vals_func):
        super().__init__(master, text="대상 데이터 고급 필터 (선택 사항)", padding=10)
        self.get_cols_func = get_cols_func
        self.get_vals_func = get_vals_func
        self.rows: list[TargetAdvFilterRow] = []
        
        self.container = ttk.Frame(self)
        self.container.pack(fill="x")
        
        btn_add = ttk.Button(self, text="+ 필터 추가", command=self.add_row)
        btn_add.pack(pady=(5, 0))
        ToolTip(btn_add, "특정 컬럼의 값으로 데이터를 필터링합니다. (예: 상태=처리완료)")

    def add_row(self):
        cols = self.get_cols_func()
        if not cols:
            messagebox.showwarning("오류", "먼저 대상 데이터를 불러와 주세요.")
            return
            
        row = TargetAdvFilterRow(self.container, cols, lambda: self.remove_row(row), self.get_vals_func)
        row.pack(fill="x", pady=2)
        self.rows.append(row)

    def remove_row(self, row):
        row.destroy()
        self.rows.remove(row)

    def get_filters(self):
        res = []
        for r in self.rows:
            f = r.get_filter()
            if f: res.append(f)
        return res

    def clear(self):
        for r in self.rows:
            r.destroy()
        self.rows = []



class App(BaseApp):
    def __init__(self, license_info=None):
        super().__init__()
        self.license_info = license_info or {"type": "personal", "expiry": "-"}
        
        # Restore missing initializations
        default_base = "file" if sys.platform == "darwin" else "open"
        default_tgt = "file"
        default_tgt = "file"
        self.base_mode = tk.StringVar(value=default_base)
        self.tgt_mode = tk.StringVar(value=default_tgt)
        
        # Output Path
        # Output Path (Default to Downloads/ExcelMatcher_Results for visibility/permission)
        default_out = os.path.join(os.path.expanduser("~"), "Downloads", "ExcelMatcher_Results")
        if not os.path.exists(default_out):
            try:
                os.makedirs(default_out, exist_ok=True)
            except:
                pass # Will be created by matcher if possible
        self.out_path = tk.StringVar(value=default_out)

        self.presets: dict[str, list[str]] = {}  # Target presets (legacy name kept for compatibility check, but we will use self.target_presets)
        self.target_presets: dict[str, list[str]] = {}
        self.base_presets: dict[str, list[str]] = {}
        self.opt_fuzzy = tk.BooleanVar(value=True)
        self.opt_color = tk.BooleanVar(value=True)
        self.opt_top10 = tk.BooleanVar(value=False)
        self.opt_match_only = tk.BooleanVar(value=False)
        self.replacer_win = None
        
        # Centralized caches for performance
        self.unique_cache = {}  # (path, sheet, header, col): values
        self.header_cache = {}  # (path, sheet, header): [cols...]
        self.sheet_cache = {}   # path: [sheets...]

        self.title(APP_TITLE)
        self.geometry("1080x1080")

        # --- HEADER SECTION (Redesigned for Commercial) ---
        # Height 100 for dramatic effect with Glassmorphism
        self.top_header = GradientFrame(self, color1="#2c3e50", color2="#1a252f", height=100) 
        self.top_header.pack(fill="x", side="top")
        
        # Draw Content on Canvas for "Translucent" Text Effect
        # 1. Logo Handling (Load Once)
        self.logo_left_tk = None
        self.logo_right_tk = None
        try:
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            logo_path = os.path.join(base_path, "assets", "logo.png")
            if os.path.exists(logo_path):
                pil_img = Image.open(logo_path)
                pil_img.thumbnail((300, 70), Image.Resampling.LANCZOS)
                self.logo_tk = ImageTk.PhotoImage(pil_img)
                
                # Split logo for "Meeting" animation
                w, h = pil_img.size
                left_half = pil_img.crop((0, 0, w//2, h))
                right_half = pil_img.crop((w//2, 0, w, h))
                self.logo_left_tk = ImageTk.PhotoImage(left_half)
                self.logo_right_tk = ImageTk.PhotoImage(right_half)
            
            # Load Gear Icon
            gear_path = os.path.join(base_path, "assets", "gear.png")
            if os.path.exists(gear_path):
                gear_img = Image.open(gear_path)
                gear_img.thumbnail((32, 32), Image.Resampling.LANCZOS)
                self.gear_tk = ImageTk.PhotoImage(gear_img)
            else:
                self.gear_tk = None
        except Exception as ex:
            print(f"Logo load error: {ex}")
            self.logo_tk = None

        # 2. Animation State
        self.assembly_progress = 0.0
        self.assembly_done = False
        self.shine_pos = -200

        def draw_header_content(event=None):
            width = self.top_header.winfo_width()
            height = self.top_header.winfo_height()
            if width < 1 or height < 1: return

            self.top_header.delete("content")
            
            y_center = height // 2
            x = 30
            
            # 1. Logo Shine (Init - only if not exists)
            if not self.top_header.find_withtag("logo_shine"):
                self.top_header.create_line(
                    -100, -100, -100, -100,
                    fill="white", width=40, capstyle="round",
                    tags=("content", "logo_shine"),
                    stipple="gray50" if sys.platform != "darwin" else "" 
                )

            # 2. Assembly Animation (E/M Matching)
            if self.logo_left_tk and self.logo_right_tk and not self.assembly_done:
                dist = 100 * (1.0 - self.assembly_progress)
                lw = self.logo_left_tk.width()
                rw = self.logo_right_tk.width()
                
                # Draw halves
                self.top_header.create_image(x - dist, y_center, image=self.logo_left_tk, anchor="w", tags="content")
                self.top_header.create_image(x + lw + dist, y_center, image=self.logo_right_tk, anchor="w", tags="content")
                
                text_x = x + lw + rw + 15
            elif self.logo_tk:
                self.top_header.create_image(x, y_center, image=self.logo_tk, anchor="w", tags="content")
                text_x = x + self.logo_tk.width() + 15
            else:
                text_x = x

            # Title: "Easy Match"
            font_title = (get_system_font()[0], 32, "bold")
            self.top_header.create_text(text_x+2, y_center-2, text="Easy Match", font=font_title, fill="#1c2833", anchor="w", tags="content")
            self.top_header.create_text(text_x, y_center-4, text="Easy Match", font=font_title, fill="white", anchor="w", tags="content")

            # 3. Gear Icon (License/Admin) on Right (Expert Position)
            gear_x = width - 95
            if hasattr(self, 'gear_tk') and self.gear_tk:
                self.top_header.create_image(gear_x, y_center, image=self.gear_tk, tags=("content", "gear_btn"))
            else:
                self.top_header.create_text(gear_x, y_center, text="S", font=(get_system_font()[0], 18, "bold"), fill="white", tags=("content", "gear_btn"))
            
            # Hover/Click effect binding for Gear
            self.top_header.tag_bind("gear_btn", "<Button-1>", lambda e: open_expert_menu(e))
            self.top_header.tag_bind("gear_btn", "<Enter>", lambda e: self.top_header.config(cursor="hand2"))
            self.top_header.tag_bind("gear_btn", "<Leave>", lambda e: self.top_header.config(cursor="arrow"))

            # 4. Help Button (?) on Right
            # Draw a circle
            btn_x = width - 50
            btn_r = 16
            
            # Circle
            # Tag 'help_btn' for binding
            self.top_header.create_oval(btn_x-btn_r, y_center-btn_r, btn_x+btn_r, y_center+btn_r, 
                                        fill="#e67e22", outline="white", width=2, tags=("content", "help_btn"))
            # Question Mark
            self.top_header.create_text(btn_x, y_center, text="?", font=(get_system_font()[0], 18, "bold"), fill="white", tags=("content", "help_btn"))
            
            # Hover/Click effect binding
            self.top_header.tag_bind("help_btn", "<Button-1>", lambda e: show_feature_info())
            self.top_header.tag_bind("help_btn", "<Enter>", lambda e: self.top_header.config(cursor="hand2"))
            self.top_header.tag_bind("help_btn", "<Leave>", lambda e: self.top_header.config(cursor="arrow"))

            # Slogan Removed as requested

        # Bind drawing to resize
        self.top_header.bind("<Configure>", draw_header_content, add="+")
        
        # Start Assembly Animation (E/M Meeting)
        def assembly_animation():
            if not self.winfo_exists(): return
            if self.assembly_done: return
            
            self.assembly_progress += 0.05
            if self.assembly_progress >= 1.0:
                self.assembly_progress = 1.0
                self.assembly_done = True
                # Start Shine Sweep after matching
                self.after(500, shine_animation)
            
            draw_header_content()
            if not self.assembly_done:
                self.after(30, assembly_animation)

        # Start Shine Sweep Animation (Modified to wait for assembly)
        def shine_animation():
            if not self.winfo_exists(): return
            if not self.assembly_done: return # Wait for pieces to meet
            
            shine_item = self.top_header.find_withtag("logo_shine")
            if not shine_item: return

            # Advance position
            self.shine_pos += 12
            if self.shine_pos > 800: # Sweep across and wait
                self.shine_pos = -200
                self.after(4000, shine_animation) # Wait 4 seconds before next sweep
                return

            # Draw diagonal shine line
            height = self.top_header.winfo_height()
            self.top_header.coords(shine_item[0],
                self.shine_pos, 0,
                self.shine_pos + 60, height
            )
            
            # Mask shine to only show over logo area (roughly x=30 to x=300)
            if 0 < self.shine_pos < 500:
                self.top_header.itemconfig(shine_item[0], state="normal")
                self.top_header.tag_lower(shine_item[0], "content")
            else:
                self.top_header.itemconfig(shine_item[0], state="hidden")

            self.after(20, shine_animation)

        # Launch Assembly first
        self.after(800, assembly_animation)


        # License/Admin Menu for Expert Gear
        def open_expert_menu(e):
            menu = tk.Menu(self, tearoff=0)
            
            def open_admin():
                from tkinter import simpledialog, messagebox
                from commercial_config import ADMIN_PASSWORD
                pw = simpledialog.askstring("관리자", "관리자 비밀번호:", show="*")
                if pw != ADMIN_PASSWORD:
                    messagebox.showerror("오류", "비밀번호가 올바르지 않습니다.")
                    return
                from ui import AdminPanel
                AdminPanel(self)

            menu.add_command(label="라이선스 등록 (제품 키 입력)", command=self.register_license)
            menu.add_command(label="관리자 패널", command=open_admin)
            menu.add_separator()
            menu.add_command(
                label=f"라이선스: {self.license_info.get('type','?')} / 만료: {self.license_info.get('expiry','?')}",
                state="disabled",
            )
            try:
                menu.tk_popup(e.x_root, e.y_root)
            finally:
                menu.grab_release()

        # Feature Info Popup (Defined here to be accessible)
        def show_feature_info():
            top = tk.Toplevel(self)
            top.title("Easy Match란?")
            top.geometry("600x500")
            top.configure(bg="white")
            
            # Center the popup
            root_x = self.winfo_rootx()
            root_y = self.winfo_rooty()
            root_w = self.winfo_width()
            root_h = self.winfo_height()
            x = root_x + (root_w // 2) - 300
            y = root_y + (root_h // 2) - 250
            top.geometry(f"600x500+{x}+{y}")

            # Title
            tk.Label(top, text="Easy Match란?", font=(get_system_font()[0], 20, "bold"), bg="white", fg="#2c3e50").pack(pady=(20, 15))
            
            # Content frame
            content = tk.Frame(top, bg="white")
            content.pack(padx=25, fill="both", expand=True)
            
            # Easy section
            easy_frame = tk.Frame(content, bg="#e8f5e9", relief="solid", borderwidth=1, highlightbackground="#27ae60", highlightthickness=1)
            easy_frame.pack(fill="x", pady=(0, 12))
            
            tk.Label(easy_frame, text="[Easy] 이지", font=(get_system_font()[0], 14, "bold"), bg="#e8f5e9", fg="#27ae60").pack(pady=(10, 5), anchor="w", padx=15)
            tk.Label(easy_frame, text="복잡한 엑셀 수식이나 코딩 없이, 누구나 쉽고 간편하게 사용할 수 있습니다.", 
                    font=(get_system_font()[0], 11), bg="#e8f5e9", fg="#2c3e50", anchor="w", wraplength=520, justify="left").pack(padx=15, pady=(0, 3), fill="x")
            tk.Label(easy_frame, text="• 숫자/문자 서식 자동 보정 기능 포함", 
                    font=(get_system_font()[0], 10), bg="#e8f5e9", fg="#34495e", anchor="w").pack(padx=30, pady=1, fill="x")
            tk.Label(easy_frame, text="• 직관적인 UI로 클릭 몇 번이면 완료", 
                    font=(get_system_font()[0], 10), bg="#e8f5e9", fg="#34495e", anchor="w").pack(padx=30, pady=(0, 10), fill="x")
            
            # Match section
            match_frame = tk.Frame(content, bg="#e3f2fd", relief="solid", borderwidth=1, highlightbackground="#2B579A", highlightthickness=1)
            match_frame.pack(fill="x", pady=(0, 12))
            
            tk.Label(match_frame, text="[Match] 매치", font=(get_system_font()[0], 14, "bold"), bg="#e3f2fd", fg="#2B579A").pack(pady=(10, 5), anchor="w", padx=15)
            tk.Label(match_frame, text="엑셀(Excel)뿐만 아니라 CSV 파일까지 복잡한 설정 없이 쉽게 매칭해줍니다.", 
                    font=(get_system_font()[0], 11), bg="#e3f2fd", fg="#2c3e50", anchor="w", wraplength=520, justify="left").pack(padx=15, pady=(0, 3), fill="x")
            tk.Label(match_frame, text="• 다양한 파일 형식 지원 (xlsx, xls, csv)", 
                    font=(get_system_font()[0], 10), bg="#e3f2fd", fg="#34495e", anchor="w").pack(padx=30, pady=1, fill="x")
            tk.Label(match_frame, text="• 자동 데이터 타입 감지", 
                    font=(get_system_font()[0], 10), bg="#e3f2fd", fg="#34495e", anchor="w").pack(padx=30, pady=(0, 10), fill="x")
            
            # Features section
            features_frame = tk.Frame(content, bg="#fff3e0", relief="solid", borderwidth=1, highlightbackground="#f39c12", highlightthickness=1)
            features_frame.pack(fill="x", pady=(0, 12))
            
            tk.Label(features_frame, text="[주요 기능]", font=(get_system_font()[0], 14, "bold"), bg="#fff3e0", fg="#e67e22").pack(pady=(10, 5), anchor="w", padx=15)
            tk.Label(features_frame, text="• 자주 쓰는 컬럼 저장 기능", 
                    font=(get_system_font()[0], 10), bg="#fff3e0", fg="#34495e", anchor="w").pack(padx=30, pady=1, fill="x")
            # Combined shorter feature list
            tk.Label(features_frame, text="• 전문가용 데이터 유무 필터 (Exists / Is Empty)", 
                    font=(get_system_font()[0], 10), bg="#fff3e0", fg="#c0392b", anchor="w").pack(padx=30, pady=1, fill="x")
            tk.Label(features_frame, text="• 초고속 로딩 및 대용량 최적화 (v1.0.19)", 
                    font=(get_system_font()[0], 10), bg="#fff3e0", fg="#34495e", anchor="w").pack(padx=30, pady=1, fill="x")
            tk.Label(features_frame, text="• 오타 자동 보정 / 치환 설정 / 색상 강조", 
                    font=(get_system_font()[0], 10), bg="#fff3e0", fg="#34495e", anchor="w").pack(padx=30, pady=(0, 10), fill="x")
            
            # Close button (Label to ensure color works on Mac)
            btn_close = tk.Label(top, text="닫기", bg="#95a5a6", fg="white", 
                      font=(get_system_font()[0], 11, "bold"), padx=25, pady=8, cursor="hand2", relief="raised")
            btn_close.pack(pady=(0, 10))
            btn_close.bind("<Button-1>", lambda e: top.destroy())
            
            # Hover effect for close button
            btn_close.bind("<Enter>", lambda e: btn_close.config(bg="#7f8c8d", relief="sunken"))
            btn_close.bind("<Leave>", lambda e: btn_close.config(bg="#95a5a6", relief="raised"))


        self._init_ui()
        
        # Apply system font
        default_font = get_system_font()
        self.option_add("*Font", default_font)
        
        self.load_presets()

        # Diagnostics will be shown on demand via toggle button

    def _init_ui(self):
        try:
            style = ttk.Style(self)
            style.theme_use("clam")
            # Modern Button Style
            style.configure("TButton", justify="center", anchor="center", font=(get_system_font()[0], 11), padding=5)
            style.configure("Header.TLabel", background="#2c3e50", foreground="white", font=(get_system_font()[0], 13, "bold"))
        except Exception:
            pass

        main = ttk.Frame(self, padding=5)  # Reduced padding for more space
        
        footer = ttk.Frame(main)
        footer.pack(side="bottom", fill="x", pady=(10, 0))

        # Collapsible Diagnostics Section
        diag_container = ttk.Frame(footer)
        diag_container.pack(side="bottom", fill="x", pady=(0, 5))
        
        # Toggle button for diagnostics
        self.diag_expanded = tk.BooleanVar(value=False)
        
        def toggle_diagnostics():
            if self.diag_expanded.get():
                diag_frame.pack_forget()
                diag_toggle_btn.config(text="▶ 환경 진단 (펼치기)")
                self.diag_expanded.set(False)
            else:
                # Show diagnostics
                diag_frame.pack(side="bottom", fill="x", pady=(5, 0))
                diag_toggle_btn.config(text="▼ 환경 진단 (접기)")
                self.diag_expanded.set(True)
                
                # Load diagnostics if not already loaded
                if self.diag_txt.get("1.0", "end-1c") == "":
                    self.diag_txt.config(state="normal")
                    self.diag_txt.insert("1.0", "환경 진단:\n")
                    self.diag_txt.insert("end", format_summary(collect_summary()))
                    self.diag_txt.config(state="disabled")
        
        diag_toggle_btn = ttk.Button(diag_container, text="▶ 환경 진단 (펼치기)", command=toggle_diagnostics)
        diag_toggle_btn.pack(side="top", fill="x")
        ToolTip(diag_toggle_btn, "시스템 환경 정보를 표시/숨김합니다\n(Python 버전, xlwings 설치 여부 등)")
        
        # Diagnostics frame (initially hidden)
        diag_frame = ttk.Frame(diag_container)
        # Don't pack initially - will be shown on toggle
        
        self.diag_txt = tk.Text(diag_frame, height=2, state="disabled", bg="#f9f9f9", fg="#555555")  # Reduced to 2
        self.diag_txt.pack(side="bottom", fill="x")

        # Collapsible Advanced Settings
        opt_container = ttk.Frame(footer)
        opt_container.pack(side="bottom", fill="x", pady=(10, 0))
        
        # Toggle button for advanced settings
        self.opt_expanded = tk.BooleanVar(value=False)
        
        def toggle_advanced():
            if self.opt_expanded.get():
                opt_frame.pack_forget()
                toggle_btn.config(text="▶ 고급 설정 (펼치기)")
                self.opt_expanded.set(False)
            else:
                opt_frame.pack(side="bottom", fill="x", pady=(5, 0))
                toggle_btn.config(text="▼ 고급 설정 (접기)")
                self.opt_expanded.set(True)
        
        toggle_btn = ttk.Button(opt_container, text="▶ 고급 설정 (펼치기)", command=toggle_advanced)
        toggle_btn.pack(side="top", fill="x")
        ToolTip(toggle_btn, "오타 보정, 치환 설정 등 고급 기능을 표시/숨김합니다")
        
        # Advanced settings frame (initially hidden)
        opt_frame = ttk.Frame(opt_container, padding=10)
        # Don't pack initially - will be shown on toggle
        
        replace_btn = ttk.Button(opt_frame, text="치환 설정 (Replace)", command=self.open_replacer)
        replace_btn.pack(side="left", padx=(0, 20))
        ToolTip(replace_btn, "데이터를 변환할 규칙을 설정합니다\n예: '남' → 'M', '여' → 'F'")
        
        fuzzy_check = ttk.Checkbutton(opt_frame, text="오타 보정 (Fuzzy Match)", variable=self.opt_fuzzy)
        fuzzy_check.pack(side="left", padx=(0, 10))
        ToolTip(fuzzy_check, "오타를 자동으로 보정합니다\n예: '홍길동' ≈ '홍길둥' (유사도 90% 이상)")
        
        color_check = ttk.Checkbutton(opt_frame, text="색상 강조 (Highlight)", variable=self.opt_color)
        color_check.pack(side="left", padx=(0, 10))
        ToolTip(color_check, "매칭된 행에 색상을 추가하여 결과를 쉽게 확인할 수 있습니다")

        top10_check = ttk.Checkbutton(opt_frame, text="전문가용: 상위 10개만 추출", variable=self.opt_top10)
        top10_check.pack(side="left", padx=(0, 10))
        ToolTip(top10_check, "대상 데이터의 상위 10개 결과만 추출합니다 (대량 데이터 샘플링용)")

        # match_only_check moved to Footer for better visibility

        # Log text area (always visible) - reduced height
        self.log_txt = tk.Text(footer, height=2, state="disabled", bg="#f0f0f0")  # Reduced to 2
        self.log_txt.pack(side="bottom", fill="x")


        # Output Path Selection
        out_frame = ttk.Frame(footer)
        out_frame.pack(side="bottom", fill="x", pady=(5, 10))
        
        ttk.Label(out_frame, text="결과 저장 위치:", font=(get_system_font()[0], 10, "bold")).pack(side="left")
        ttk.Entry(out_frame, textvariable=self.out_path, state="readonly").pack(side="left", fill="x", expand=True, padx=5)
        
        def pick_out_dir():
            d = filedialog.askdirectory()
            if d: self.out_path.set(d)
            
        ttk.Button(out_frame, text="폴더 변경", command=pick_out_dir).pack(side="left")

        # Option: Match Only (Moved from Advanced to here for visibility)
        match_only_frame = ttk.Frame(footer)
        match_only_frame.pack(side="bottom", fill="x", pady=(0, 5))
        
        # Use a style to make it stand out or just standard
        mo_chk = ttk.Checkbutton(match_only_frame, text="[중요] 매칭된 결과만 저장 (미매칭 제외)", variable=self.opt_match_only)
        mo_chk.pack(anchor="center")
        ToolTip(mo_chk, "체크 시: 매칭에 성공한 행만 저장합니다.\n해제 시: 원본 기준 데이터의 모든 행을 유지하며 매칭된 정보를 붙입니다.")

        # Run button with Green styling (Matching Usage Guide)
        # Run button with Green styling (Matching Usage Guide) - Using Label for macOS color support
        run_btn = tk.Label(
            footer, 
            text="매칭 실행 (RUN)", 
            bg="#16a085",  # Green/Teal
            fg="#ffffff",  # White
            font=(get_system_font()[0], 15, "bold"),
            padx=20,
            pady=15,
            cursor="hand2",
            relief="raised",
            borderwidth=2
        )
        # Use simple Button binding instead of Label if possible, but keep Label for color
        # Re-verify Run button to ensure it uses Label as previously fixed.
        # Wait, step 3024 replaced Inquiry button to Label, but did not touch Run button (except maybe context).
        # Run button was already Label in Step 2981.
        # I just need to make sure I don't break anything.
        # This chunk is just context, no replacement needed here.
        # But I will confirm Run Button is distinct from preset logic.
        pass

        run_btn.pack(side="bottom", fill="x", pady=(10, 5))
        ToolTip(run_btn, "설정한 조건으로 데이터 매칭을 시작합니다\n(단축키: Ctrl+M)\n결과는 설정된 저장 위치에 생성됩니다")
        
        # Hover effect for run button
        def on_run_enter(e):
            run_btn.config(bg="#1abc9c", relief="sunken") # Lighter green
        
        def on_run_leave(e):
            run_btn.config(bg="#16a085", relief="raised")
            
        def on_run_click(e):
            self.run()
        
        run_btn.bind("<Enter>", on_run_enter)
        run_btn.bind("<Leave>", on_run_leave)
        run_btn.bind("<Button-1>", on_run_click)


        # --- 2. File Loaders (Top - Split Layout) ---
        top_content = ttk.Frame(main)
        top_content.pack(side="top", fill="x", pady=(0, 10))

        # Split Container for Files
        split_frame = ttk.Frame(top_content)
        split_frame.pack(fill="x", expand=True)

        # Left: Source Loader
        self.src_loader = FileLoaderFrame(split_frame, "1. 기준 데이터 (Key 보유)", self.base_mode, on_change=self._load_base_cols, on_fetch_vals=self._fetch_base_unique_vals)
        self.src_loader.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Right: Target Loader
        self.tgt_loader = FileLoaderFrame(split_frame, "2. 대상 데이터 (데이터 가져올 곳)", self.tgt_mode, on_change=self._load_tgt_cols, on_fetch_vals=self._fetch_tgt_unique_vals)
        self.tgt_loader.pack(side="right", fill="both", expand=True, padx=(5, 0))

        # Advanced Filter Setup
        self.filter_win = None
        self.tgt_filter_ui = TargetFilterFrame(None, self._get_tgt_cols, self._fetch_tgt_unique_vals)
        
        btn_adv_filter = ttk.Button(split_frame, text="대상 데이터 고급 필터 설정", command=self.open_advanced_filter)
        btn_adv_filter.pack(side="bottom", anchor="e", pady=(5, 0))
        ToolTip(btn_adv_filter, "대상 데이터에 대해 상세 필터(여러 값 동시 선택 등)를 설정합니다.")

        # --- 3. Column Selectors (Middle - Split Layout) ---
        col_content = ttk.Frame(main)
        col_content.pack(side="top", fill="both", expand=True, pady=0)

        # Left Container (Base Data)
        left_panel = ttk.Frame(col_content)
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Base Presets Bar
        base_preset_frame = ttk.Frame(left_panel)
        base_preset_frame.pack(side="top", fill="x", pady=(0, 5))
        
        ttk.Label(base_preset_frame, text="자주 쓰는 컬럼:", font=get_system_font()).pack(side="left")

        self.cb_preset_base = ttk.Combobox(base_preset_frame, state="readonly", width=15)
        self.cb_preset_base.pack(side="left", padx=5)
        self.cb_preset_base.bind("<<ComboboxSelected>>", self.apply_base_preset)
        ToolTip(self.cb_preset_base, "저장된 기준 키 설정을 불러옵니다")
        
        ttk.Button(base_preset_frame, text="저장", command=self.save_base_preset, width=4).pack(side="left", padx=2)
        ttk.Button(base_preset_frame, text="삭제", command=self.delete_base_preset, width=4).pack(side="left", padx=2)
        
        # Save Conditions to Excel Button
        btn_save_sheet = ttk.Button(base_preset_frame, text="설정파일 저장", command=self.save_conditions_to_sheet, width=10)
        btn_save_sheet.pack(side="right", padx=2)
        ToolTip(btn_save_sheet, "현재 설정된 조건(키, 대상 컬럼)을 기준 데이터 파일에 새로운 시트로 저장합니다")

        # Left: Match Key Selector
        self.match_key_selector = ColumnSelectorFrame(left_panel, "매칭 키 (Key) 선택 - 기준 데이터")
        self.match_key_selector.pack(side="bottom", fill="both", expand=True)

        # Right Container (Presets + Target Columns)
        right_panel = ttk.Frame(col_content)
        right_panel.pack(side="right", fill="both", expand=True, padx=(5, 0))

        # Presets Bar (Moved here)
        preset_frame = ttk.Frame(right_panel)
        preset_frame.pack(side="top", fill="x", pady=(0, 5))
        
        ttk.Label(preset_frame, text="자주 쓰는 컬럼:", font=get_system_font()).pack(side="left")

        self.cb_preset = ttk.Combobox(preset_frame, state="readonly", width=18)
        self.cb_preset.pack(side="left", padx=5)
        self.cb_preset.bind("<<ComboboxSelected>>", self.apply_preset)
        ToolTip(self.cb_preset, "저장된 컬럼 설정을 불러옵니다")
        
        ttk.Button(preset_frame, text="저장", command=self.save_preset, width=5).pack(side="left", padx=2)
        ttk.Button(preset_frame, text="삭제", command=self.delete_preset, width=5).pack(side="left", padx=2)

        # Right: Target Column Selector
        self.target_col_selector = ColumnSelectorFrame(right_panel, "가져올 컬럼 선택 - 대상 데이터")
        self.target_col_selector.pack(side="bottom", fill="both", expand=True)
        
        # Force menu update
        def force_menu():
            try:
                if hasattr(self, 'menubar'):
                    self.config(menu=self.menubar)
                self.update_idletasks()
            except:
                pass
        self.after(500, force_menu)

        # --- COMMERCIAL FOOTER (pack BEFORE main to stay at bottom) ---
        from commercial_config import CREATOR_NAME, SALES_INFO, PRICE_INFO, CONTACT_INFO, TRIAL_EMAIL, TRIAL_SUBJECT, SUPPORT_MESSAGE, DONATION_CTA
        c_footer = tk.Frame(self, bg="#2c3e50")
        c_footer.pack(side="bottom", fill="x", pady=0)
        c_footer.config(height=60)
        
        # Left: Creator name (Updated for Commercial Release)
        tk.Label(
            c_footer,
            text=f"이지매치 (EasyMatch)\nDeveloped by {CREATOR_NAME}",
            bg="#2c3e50",
            fg="#bdc3c7",  # Slightly brighter than #95a5a6 for readability
            font=(get_system_font()[0], 10),
            justify="left"
        ).pack(side="left", padx=20, pady=10)
        
        # Comprehensive inquiry popup
        def show_inquiry_popup(e=None):
            top = tk.Toplevel(self)
            top.title("문의")
            top.geometry("450x520")
            top.configure(bg="white")
            top.resizable(False, False)
            
            # Center
            top.update_idletasks()
            x = (top.winfo_screenwidth() // 2) - 225
            y = (top.winfo_screenheight() // 2) - 260
            top.geometry(f"450x520+{x}+{y}")
            
            # Main container
            container = tk.Frame(top, bg="white")
            container.pack(fill="both", expand=True, padx=25, pady=20)
            
            # Title (centered)
            tk.Label(
                container,
                text="문의",
                font=(get_system_font()[0], 16, "bold"),
                bg="white",
                fg="#2c3e50"
            ).pack(pady=(0, 20), anchor="center")
            
            # Section 1: Pricing
            price_frame = tk.Frame(container, bg="#ecf0f1", relief="solid", borderwidth=1)
            price_frame.pack(fill="x", pady=(0, 12))
            
            tk.Label(price_frame, text="> 가격 안내", font=(get_system_font()[0], 12, "bold"), bg="#ecf0f1", fg="#2c3e50").pack(anchor="w", padx=12, pady=(10, 4))
            tk.Label(price_frame, text="개인(1PC): 1년 3.3만 / 평생 13.2만 (최대 100만행)", font=(get_system_font()[0], 9), bg="#ecf0f1", fg="#34495e", anchor="w").pack(anchor="w", padx=15, pady=1)
            tk.Label(price_frame, text="기업(1PC): 영구 180,000원 (무제한)", font=(get_system_font()[0], 9, "bold"), bg="#ecf0f1", fg="#c0392b", anchor="w").pack(anchor="w", padx=15, pady=(1, 10))
            
            # Section 2: Payment/Donation Account
            payment_frame = tk.Frame(container, bg="#e8f5e9", relief="solid", borderwidth=1)
            payment_frame.pack(fill="x", pady=(0, 12))
            
            tk.Label(payment_frame, text="> 입금(후원) 계좌", font=(get_system_font()[0], 12, "bold"), bg="#e8f5e9", fg="#2c3e50").pack(anchor="w", padx=12, pady=(10, 4))
            tk.Label(payment_frame, text="카카오뱅크 3333-03-9648364 (박희본)", font=(get_system_font()[0], 10, "bold"), bg="#e8f5e9", fg="#c0392b", anchor="w").pack(anchor="w", padx=15, pady=2)
            
            def copy_account():
                top.clipboard_clear()
                top.clipboard_append("3333-03-9648364")
                top.update()
                messagebox.showinfo("완료", "계좌번호가 복사되었습니다!", parent=top)
            
            # Copy Button (Label for macOS color support)
            btn_copy_acc = tk.Label(payment_frame, text="계좌번호 복사", bg="#7f8c8d", fg="white", 
                         font=(get_system_font()[0], 9, "bold"), padx=12, pady=4, cursor="hand2", relief="raised")
            btn_copy_acc.pack(anchor="w", padx=15, pady=(4, 10))
            btn_copy_acc.bind("<Button-1>", lambda e: copy_account())
            # Hover
            btn_copy_acc.bind("<Enter>", lambda e: btn_copy_acc.config(bg="#5d6d6e", relief="sunken"))
            btn_copy_acc.bind("<Leave>", lambda e: btn_copy_acc.config(bg="#7f8c8d", relief="raised"))

            
            # Section 3: Customization Contact
            contact_frame = tk.Frame(container, bg="#fff3e0", relief="solid", borderwidth=1)
            contact_frame.pack(fill="x", pady=(0, 20))
            
            tk.Label(contact_frame, text="> 커스터마이징 문의", font=(get_system_font()[0], 12, "bold"), bg="#fff3e0", fg="#2c3e50").pack(anchor="w", padx=12, pady=(10, 4))
            tk.Label(contact_frame, text="라이선스 구매/커스터마이징: bough38@gmail.com", font=(get_system_font()[0], 10, "bold"), bg="#fff3e0", fg="#c0392b", anchor="w").pack(anchor="w", padx=15, pady=1)
            tk.Label(contact_frame, text="디스코드: bough38 (세은아빠)", font=(get_system_font()[0], 10, "bold"), bg="#fff3e0", fg="#2980b9", anchor="w").pack(anchor="w", padx=15, pady=(1, 10))
            
            def copy_email():
                top.clipboard_clear()
                top.clipboard_append("bough38@gmail.com")
                top.update()
                messagebox.showinfo("완료", "이메일 주소가 복사되었습니다!", parent=top)
            
            # Copy Email Button
            btn_copy_email = tk.Label(contact_frame, text="이메일 복사", bg="#7f8c8d", fg="white", 
                           font=(get_system_font()[0], 9, "bold"), padx=12, pady=4, cursor="hand2", relief="raised")
            btn_copy_email.pack(anchor="w", padx=15, pady=(0, 10))
            btn_copy_email.bind("<Button-1>", lambda e: copy_email())
            # Hover
            btn_copy_email.bind("<Enter>", lambda e: btn_copy_email.config(bg="#5d6d6e", relief="sunken"))
            btn_copy_email.bind("<Leave>", lambda e: btn_copy_email.config(bg="#7f8c8d", relief="raised"))

            
            # Close button (centered, larger) - Label for macOS
            btn_close_popup = tk.Label(container, text="닫기", bg="#95a5a6", fg="white", 
                            font=(get_system_font()[0], 11, "bold"), padx=30, pady=8, cursor="hand2", relief="raised")
            btn_close_popup.pack(pady=(10, 0), anchor="center")
            btn_close_popup.bind("<Button-1>", lambda e: top.destroy())
            # Hover
            btn_close_popup.bind("<Enter>", lambda e: btn_close_popup.config(bg="#7f8c8d", relief="sunken"))
            btn_close_popup.bind("<Leave>", lambda e: btn_close_popup.config(bg="#95a5a6", relief="raised"))

        
        # Right: Inquiry button (Footer) - Label for macOS color support
        inquiry_btn = tk.Label(
            c_footer,
            text="문의",
            bg="#3498db",
            fg="white",
            font=(get_system_font()[0], 11, "bold"),
            padx=20,
            pady=8,
            cursor="hand2",
            relief="raised",
            borderwidth=2
        )
        inquiry_btn.pack(side="right", padx=20, pady=10)
        inquiry_btn.bind("<Button-1>", show_inquiry_popup)
        
        def on_enter_inquiry(e):
            inquiry_btn.config(bg="#2980b9", relief="sunken")
        def on_leave_inquiry(e):
            inquiry_btn.config(bg="#3498db", relief="raised")
        
        inquiry_btn.bind("<Enter>", on_enter_inquiry)
        inquiry_btn.bind("<Leave>", on_leave_inquiry)

        # --- Guide Button (Next to Inquiry) ---
        def open_usage_guide_click(e=None):
            import webbrowser
            try:
                # Handle path for frozen app (PyInstaller)
                if getattr(sys, 'frozen', False):
                    base_path = sys._MEIPASS
                else:
                    base_path = os.path.dirname(os.path.abspath(__file__))
                
                guide_path = os.path.join(base_path, "usage_guide.html")
                
                if not os.path.exists(guide_path):
                     guide_path = os.path.abspath("usage_guide.html")

                webbrowser.open(f"file://{guide_path}")
            except Exception as ex:
                messagebox.showinfo("안내", f"사용 가이드 (usage_guide.html) 파일을 찾을 수 없습니다.\n{ex}")

        guide_btn = tk.Label(
            c_footer,
            text="사용방법",
            bg="#16a085",
            fg="white",
            font=(get_system_font()[0], 11, "bold"),
            padx=15,
            pady=8,
            cursor="hand2",
            relief="raised",
            borderwidth=2
        )
        guide_btn.pack(side="right", padx=(0, 0), pady=10)
        guide_btn.bind("<Button-1>", open_usage_guide_click)
        
        def on_enter_guide(e): guide_btn.config(bg="#138d75", relief="sunken")
        def on_leave_guide(e): guide_btn.config(bg="#16a085", relief="raised")
        guide_btn.bind("<Enter>", on_enter_guide)
        guide_btn.bind("<Leave>", on_leave_guide)

        
        # Pack main frame AFTER footer
        main.pack(fill="both", expand=True)

        # 초기 로드
        self._load_base_cols()
        self._load_tgt_cols()

        # Monitoring: Report Usage Status to Discord
        try:
            from monitor import report_usage_status
            from commercial_config import DISCORD_WEBHOOK_URL
            report_usage_status(DISCORD_WEBHOOK_URL)
        except Exception as e:
            import logging
            logging.error(f"Failed to initiate monitoring: {e}")

        # macOS에서 앱이 백그라운드에 가려지는 문제 방지 (앞으로 가져오기)
        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))
        self.focus_force()

    def _log(self, msg: str):
        self.log_txt.config(state="normal")
        self.log_txt.insert("end", f"- {msg}\n")
        self.log_txt.see("end")
        self.log_txt.config(state="disabled")
        self.update_idletasks()

    def open_advanced_filter(self):
        """Open advanced filter settings in a popup window"""
        if self.filter_win is None or not self.filter_win.winfo_exists():
            self.filter_win = tk.Toplevel(self)
            self.filter_win.title("대상 데이터 고급 필터 설정")
            self.filter_win.geometry("600x400")
            self.filter_win.transient(self)
            
            # Re-parent or re-create the filter UI inside the popup
            container = ttk.Frame(self.filter_win, padding=20)
            container.pack(fill="both", expand=True)
            
            # Clear and Re-create for safety on re-open
            for child in container.winfo_children(): child.destroy()
            
            self.tgt_filter_ui = TargetFilterFrame(container, self._get_tgt_cols, self._fetch_tgt_unique_vals)
            self.tgt_filter_ui.pack(fill="both", expand=True)
            
            # Add a close button at bottom
            ttk.Button(container, text="확인 (적용)", command=self.filter_win.destroy).pack(pady=(10, 0))
        else:
            self.filter_win.lift()

    def open_replacer(self):
        if self.replacer_win is None or not self.replacer_win.winfo_exists():
            self.replacer_win = ReplacementEditor(self)
        else:
            self.replacer_win.lift()

    def register_license(self):
        from license_manager import validate_key, save_license_key
        from security_utils import get_hwid
        
        # 1. Input Key
        hwid = get_hwid()
        key = simpledialog.askstring("라이선스 등록", 
            f"제품 키를 입력하세요:\n(현재 기기 ID: {hwid})\n\n키 형식: EM-XXXX-XXXX...", 
            parent=self)
        if not key:
            return
            
        key = key.strip()
        if not key:
            return

        # 2. Validate
        valid, info = validate_key(key)
        if not valid:
            messagebox.showerror("등록 실패", "유효하지 않은 라이선스 키입니다.")
            return

        # 3. Save
        try:
            save_license_key(key)
            self.license_info = info
            
            # 4. Update UI (Menu)
            # Re-create menu or just show success message (Menu update is tricky dynamically without refactor)
            # For now, just show success and tell to restart for full effect if needed, 
            # though current runtime specific limits might not update until restart depending on logic.
            # But we can update the license_info dict which is used in run().
            
            messagebox.showinfo("등록 성공", f"라이선스가 등록되었습니다.\n타입: {info.get('type')}\n만료: {info.get('expiry')}")
            
        except Exception as e:
            messagebox.showerror("오류", f"라이선스 저장 중 오류가 발생했습니다.\n{e}")

    # ----------------
    # Preset (columns)
    # ----------------
    def load_presets(self):
        try:
            if os.path.exists(PRESET_FILE):
                with open(PRESET_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
                    
                    # Check structure (Legacy vs New)
                    if "base" in data or "target" in data:
                        self.base_presets = data.get("base", {}) or {}
                        self.target_presets = data.get("target", {}) or {}
                    else:
                        # Legacy format (flat dict -> target presets)
                        self.base_presets = {}
                        self.target_presets = data
            else:
                self.base_presets = {}
                self.target_presets = {}
        except Exception:
            self.base_presets = {}
            self.target_presets = {}

        # Update Target CB
        self.cb_preset["values"] = list(self.target_presets.keys())
        self.cb_preset.set("설정을 선택하세요" if self.target_presets else "저장된 설정 없음")
        
        # Update Base CB
        self.cb_preset_base["values"] = list(self.base_presets.keys())
        self.cb_preset_base.set("설정을 선택하세요" if self.base_presets else "저장된 설정 없음")

    def _save_presets(self):
        with open(PRESET_FILE, "w", encoding="utf-8") as f:
            data = {
                "base": self.base_presets,
                "target": self.target_presets
            }
            json.dump(data, f, ensure_ascii=False, indent=4)

    # --- Target Presets ---
    def save_preset(self):
        items = self.target_col_selector.get_selected()
        if not items:
            messagebox.showwarning("경고", "저장할 컬럼을 하나 이상 선택하세요.")
            return
        name = simpledialog.askstring("설정 저장", "이 설정의 이름을 입력하세요:\n(예: 급여대장용)")
        if not name: return
        name = name.strip()
        if not name: return
        self.target_presets[name] = items
        self._save_presets()
        self.load_presets()
        self.cb_preset.set(name)
        messagebox.showinfo("저장 완료", f"[{name}] 설정이 저장되었습니다.")

    def delete_preset(self):
        name = self.cb_preset.get()
        if name in self.target_presets and messagebox.askyesno("삭제 확인", f"정말 [{name}] 설정을 삭제하시겠습니까?"):
            del self.target_presets[name]
            self._save_presets()
            self.load_presets()
            self.cb_preset.set("")

    def apply_preset(self, event=None):
        name = self.cb_preset.get()
        if name in self.target_presets:
            items = self.target_presets[name]
            cnt = self.target_col_selector.list.set_checked_items(items)
            self._log(f"대상 프리셋 [{name}] 적용됨 ({cnt}개 항목 선택)")

    # --- Base Presets ---
    def save_base_preset(self):
        items = self.match_key_selector.get_selected()
        if not items:
            messagebox.showwarning("경고", "저장할 키 컬럼을 하나 이상 선택하세요.")
            return
        name = simpledialog.askstring("키 설정 저장", "이 키 설정의 이름을 입력하세요:\n(예: 주민번호기준)")
        if not name: return
        name = name.strip()
        if not name: return
        self.base_presets[name] = items
        self._save_presets()
        self.load_presets()
        self.cb_preset_base.set(name)
        messagebox.showinfo("저장 완료", f"[{name}] 키 설정이 저장되었습니다.")

    def delete_base_preset(self):
        name = self.cb_preset_base.get()
        if name in self.base_presets and messagebox.askyesno("삭제 확인", f"정말 [{name}] 키 설정을 삭제하시겠습니까?"):
            del self.base_presets[name]
            self._save_presets()
            self.load_presets()
            self.cb_preset_base.set("")

    def apply_base_preset(self, event=None):
        name = self.cb_preset_base.get()
        if name in self.base_presets:
            items = self.base_presets[name]
            cnt = self.match_key_selector.list.set_checked_items(items)
            self._log(f"기준 프리셋 [{name}] 적용됨 ({cnt}개 항목 선택)")
            
    # --- Save Conditions to Excel ---
    def save_conditions_to_sheet(self):
        """Save current selection (keys, targets) to a new sheet in Base Data file"""
        b_cfg = self.src_loader.get_config()
        if not b_cfg["path"] and not b_cfg["book"]:
            messagebox.showwarning("경고", "기준 데이터가 로드되지 않았습니다.")
            return

        keys = self.match_key_selector.get_selected()
        targets = self.target_col_selector.get_selected()
        
        if not keys and not targets:
            messagebox.showwarning("경고", "저장할 설정(키 또는 대상 컬럼)이 없습니다.")
            return
            
        try:
            from openpyxl import load_workbook, Workbook
            import pandas as pd
            import datetime
            
            # Prepare data
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data_row = [
                timestamp,
                ",".join(keys),
                ",".join(targets),
                str(self.opt_fuzzy.get()),
                str(self.opt_color.get())
            ]
            header = ["Timestamp", "Match Keys", "Target Columns", "Fuzzy Option", "Color Option"]
            sheet_name = "EasyMatch_Conditions"

            # Function to append data using pandas for simplicity if possible, or openpyxl
            # Since we want to append to existing file or create new sheet.
            
            file_path = b_cfg["path"] if b_cfg["type"] == "file" else None
            
            if file_path and os.path.exists(file_path):
                # File mode
                try:
                    wb = load_workbook(file_path)
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(sheet_name)
                        ws.append(header)
                    else:
                        ws = wb[sheet_name]
                        
                    ws.append(data_row)
                    wb.save(file_path)
                    messagebox.showinfo("저장 완료", f"기준 데이터 파일에 [{sheet_name}] 시트가 추가/업데이트 되었습니다.")
                except Exception as e:
                    messagebox.showerror("저장 실패", f"파일 저장 중 오류: {e}\n(파일이 열려있다면 닫고 시도하세요)")
            
            elif b_cfg["type"] == "open" and b_cfg["book"]:
                # Open Excel mode (xlwings)
                try:
                    import xlwings as xw
                    app = xw.apps.active
                    wb = None
                    try:
                        wb = xw.books[b_cfg["book"]]
                    except:
                        # Try to find by name
                        for b in xw.books:
                            if b.name == b_cfg["book"]:
                                wb = b
                                break
                    if not wb:
                        raise Exception("엑셀 파일을 찾을 수 없습니다.")
                        
                    # Check sheet
                    sheet_exists = False
                    for s in wb.sheets:
                        if s.name == sheet_name:
                            sheet_exists = True
                            break
                    
                    if not sheet_exists:
                        ws = wb.sheets.add(sheet_name)
                        ws.range("A1").value = header
                    else:
                        ws = wb.sheets[sheet_name]
                    
                    # Find next empty row
                    # Simple approach: count used range
                    last_row = ws.range(f"A{ws.cells.last_cell.row}").end('up').row
                    ws.range(f"A{last_row+1}").value = data_row
                    
                    messagebox.showinfo("저장 완료", f"활성 엑셀 파일에 [{sheet_name}] 시트가 업데이트 되었습니다.")
                    
                except Exception as e:
                    messagebox.showerror("저장 실패", f"열려있는 엑셀 제어 실패: {e}")
            else:
                 messagebox.showwarning("오류", "파일을 찾을 수 없거나 저장할 수 없는 상태입니다.")

        except Exception as e:
            messagebox.showerror("오류", f"저장 중 예기치 않은 오류 발생: {e}")
    
    # -------------
    # Header loaders
    # -------------
    def _fetch_headers(self, path, sheet, header):
        key = (path, sheet, header)
        if key in self.header_cache: return self.header_cache[key]
        try:
            from excel_io import read_header_file
            cols = read_header_file(path, sheet, header)
            if cols: self.header_cache[key] = cols
            return cols
        except: return []

    def _fetch_sheet_names(self, path):
        if path in self.sheet_cache: return self.sheet_cache[path]
        try:
            from excel_io import get_sheet_names
            sheets = get_sheet_names(path)
            if sheets: self.sheet_cache[path] = sheets
            return sheets
        except: return []

    def _load_base_cols(self):
        if not hasattr(self, 'src_loader'): return
        cfg = self.src_loader.get_config()
        if not cfg.get("sheet"): return
        try:
            if cfg["type"] == "file":
                if not cfg["path"] or not os.path.exists(cfg["path"]): return
                cols = self._fetch_headers(cfg["path"], cfg["sheet"], cfg["header"])
            else:
                if not cfg["book"]: return
                cols = read_header_open(cfg["book"], cfg["sheet"], cfg["header"])
            
            self.match_key_selector.set_items(cols)
        except Exception as e:
            self._log(f"기준 헤더 로드 실패: {e}")

    def _load_tgt_cols(self):
        if not hasattr(self, 'tgt_loader'): return
        cfg = self.tgt_loader.get_config()
        if cfg["type"] == "file" and not cfg["path"]: return
        if cfg["type"] == "open" and not cfg["book"]: return
        try:
            if cfg["type"] == "file":
                cols = self._fetch_headers(cfg["path"], cfg["sheet"], cfg["header"])
            else:
                cols = read_header_open(cfg["book"], cfg["sheet"], cfg["header"])
            self.target_col_selector.set_items(cols)
            self._log(f"대상 컬럼 로드됨 ({len(cols)}개)")
        except Exception as e:
            self._log(f"대상 헤더 로드 실패: {e}")

    def _get_tgt_cols(self):
        cfg = self.tgt_loader.get_config()
        try:
            if cfg["type"] == "file" and cfg["path"]:
                return self._fetch_headers(cfg["path"], cfg["sheet"], cfg["header"])
            elif cfg["type"] == "open" and cfg["book"]:
                return read_header_open(cfg["book"], cfg["sheet"], cfg["header"])
        except: pass
        return []

    def _fetch_base_unique_vals(self, col):
        from excel_io import get_unique_values
        cfg = self.src_loader.get_config()
        if cfg["type"] == "file" and cfg["path"]:
            # Check cache
            cache_key = (cfg["path"], cfg["sheet"], cfg["header"], col)
            if cache_key in self.unique_cache:
                return self.unique_cache[cache_key]
            
            try:
                vals = get_unique_values(cfg["path"], cfg["sheet"], cfg["header"], col)
                if vals:
                    self.unique_cache[cache_key] = vals
                return vals
            except: pass
        elif cfg["type"] == "open" and cfg["book"]:
            from open_excel import read_table_open
            try:
                df = read_table_open(cfg["book"], cfg["sheet"], cfg["header"], [col])
                if not df.empty:
                    return sorted(df[col].unique().tolist())
            except: pass
        return []

    def _fetch_tgt_unique_vals(self, col):
        from excel_io import get_unique_values
        cfg = self.tgt_loader.get_config()
        if cfg["type"] == "file" and cfg["path"]:
            # Check cache
            cache_key = (cfg["path"], cfg["sheet"], cfg["header"], col)
            if cache_key in self.unique_cache:
                return self.unique_cache[cache_key]
                
            try:
                vals = get_unique_values(cfg["path"], cfg["sheet"], cfg["header"], col)
                if vals:
                    self.unique_cache[cache_key] = vals
                return vals
            except: pass
        elif cfg["type"] == "open" and cfg["book"]:
            from open_excel import read_table_open
            try:
                df = read_table_open(cfg["book"], cfg["sheet"], cfg["header"], [col])
                if not df.empty:
                    return sorted(df[col].unique().tolist())
            except: pass
        return []

    def _clear_unique_cache(self):
        """Clears all caches (called when file path, sheet, or header change)"""
        self.unique_cache = {}
        self.header_cache = {}
        self.sheet_cache = {}
        self._log("데이터 캐시가 초기화되었습니다.")

    # ----
    # Run
    # ----
    def run(self):
        """Execute matching with progress dialog"""
        try:
            # Validate inputs
            b_cfg = self.src_loader.get_config()
            t_cfg = self.tgt_loader.get_config()
            keys = self.match_key_selector.get_selected()
            take = self.target_col_selector.get_selected()

            if not keys:
                messagebox.showwarning("경고", "매칭할 키(Key)를 하나 이상 선택하세요.")
                return
            if not take:
                messagebox.showwarning("경고", "가져올 컬럼을 선택하세요.")
                return

            options = {
                "fuzzy": self.opt_fuzzy.get(),
                "color": self.opt_color.get(),
                "top10": self.opt_top10.get(),
                "match_only": self.opt_match_only.get()
            }
            
            # Replacement Rules
            active_replace_rules = _load_replace_file()["active"]

            # Filters
            applied_filters = {}
            
            src_f = self.src_loader.get_filters()
            if src_f:
                applied_filters["base_multi"] = src_f
            
            tgt_f = self.tgt_loader.get_filters()
            if tgt_f:
                applied_filters["target_multi"] = tgt_f
            
            target_adv = self.tgt_filter_ui.get_filters()
            if target_adv:
                applied_filters["target_advanced"] = target_adv

            # Create progress dialog
            self._show_progress_dialog(b_cfg, t_cfg, keys, take, options, active_replace_rules, applied_filters)

        except Exception as e:
            traceback.print_exc()
            msg = str(e)

            if "xlwings" in msg.lower():
                msg += (
                    "\n\n[힌트]\n"
                    "- Excel 설치 확인\n"
                    "- macOS: 개인정보 보호 및 보안 > 자동화에서 Terminal/iTerm2의 Excel 제어 허용\n"
                    "- 파일 모드로 사용 가능"
                )

            messagebox.showerror("오류", f"실행 중 오류:\n{msg}")
            self._log(f"Error: {msg}")

    def _show_progress_dialog(self, b_cfg, t_cfg, keys, take, options, active_replace_rules, filters):
        """Show progress dialog and run matching in background thread"""
        import threading
        
        # Create progress window
        progress_win = tk.Toplevel(self)
        progress_win.title("매칭 진행 중...")
        progress_win.geometry("500x200")
        progress_win.resizable(False, False)
        progress_win.transient(self)
        progress_win.grab_set()
        
        # Center the window
        progress_win.update_idletasks()
        x = (progress_win.winfo_screenwidth() // 2) - (500 // 2)
        y = (progress_win.winfo_screenheight() // 2) - (240 // 2)
        progress_win.geometry(f"500x240+{x}+{y}")
        
        # Progress frame
        frame = tk.Frame(progress_win, bg="white", padx=30, pady=30)
        frame.pack(fill="both", expand=True)
        
        # Title
        tk.Label(
            frame,
            text="데이터 매칭 중...",
            font=(get_system_font()[0], 14, "bold"),
            bg="white",
            fg="#2c3e50"
        ).pack(pady=(0, 15))
        
        # Progress bar
        progress_bar = ttk.Progressbar(
            frame,
            mode='determinate',
            length=400,
            maximum=100
        )
        progress_bar.pack(pady=(0, 10))
        
        # Status label
        status_label = tk.Label(
            frame,
            text="준비 중...",
            font=(get_system_font()[0], 11),
            bg="white",
            fg="#7f8c8d"
        )
        status_label.pack(pady=(0, 15))
        
        # Cancel button
        cancel_flag = {"cancelled": False}
        
        def cancel_matching(e=None):
            if not cancel_flag["cancelled"]:
                cancel_flag["cancelled"] = True
                cancel_btn.config(text="취소 중...", bg="#95a5a6", cursor="arrow")
        
        cancel_btn = tk.Label(
            frame,
            text="종료 및 취소",
            bg="#e74c3c",
            fg="white",
            font=(get_system_font()[0], 12, "bold"),
            padx=30,
            pady=20,
            cursor="hand2",
            relief="flat"
        )
        cancel_btn.pack(pady=15)
        
        # Bind interactions for the custom label-button
        cancel_btn.bind("<Button-1>", cancel_matching)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#c0392b") if not cancel_flag["cancelled"] else None)
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#e74c3c") if not cancel_flag["cancelled"] else None)
        
        # Result storage
        result = {"out_path": None, "summary": None, "error": None}
        
        def update_progress(value, message):
            """Update progress from worker thread (thread-safe)"""
            def _update():
                if not progress_win.winfo_exists():
                    return
                if value is not None:
                    progress_bar['value'] = value
                status_label.config(text=message)
                progress_win.update_idletasks()
            
            # Schedule update in main thread
            self.after(0, _update)
        
        def worker_thread():
            """Background worker thread for matching"""
            # Local Debug Logger
            def _log_ui(msg):
                try:
                    import datetime
                    log_path = os.path.join(os.path.expanduser("~"), "Desktop", "EasyMatch_Log.txt")
                    with open(log_path, "a", encoding="utf-8") as f:
                        f.write(f"[{datetime.datetime.now()}] [UI-Thread] {msg}\n")
                except: pass

            _log_ui("Thread Started")
            try:
                # Initialize COM for Windows threading compatibility (just in case)
                try:
                    import pythoncom
                    pythoncom.CoInitialize()
                except:
                    pass

                if cancel_flag["cancelled"]:
                    _log_ui("Cancelled at start")
                    return
                
                update_progress(10, "파일 읽는 중...")
                
                # Perform matching with progress updates
                output_dir = self.out_path.get()
                if not output_dir: 
                    _log_ui("Output Dir empty. Fallback.")
                    output_dir = OUT_DIR # fallback (but we set default in init)
                
                _log_ui(f"Calling match_universal. Out: {output_dir}")
                
                out_path, summary, preview = match_universal(
                    b_cfg, t_cfg, keys, take, output_dir, options, active_replace_rules, filters,
                    lambda msg, val=None: update_progress(val, msg),
                    cancel_check=lambda: cancel_flag["cancelled"]
                )
                
                if cancel_flag["cancelled"]:
                    _log_ui("Cancelled after match")
                    return
                
                update_progress(95, "결과 저장 중...")
                result["out_path"] = out_path
                result["summary"] = summary
                
                _log_ui(f"Success. OutPath: {out_path}")
                update_progress(100, "완료!")
                
                # Close progress window after short delay
                def close_and_show_result():
                    if progress_win.winfo_exists():
                        progress_win.destroy()
                    
                    if not cancel_flag["cancelled"]:
                        msg = (
                            "작업 완료!\n\n"
                            "[결과 리포트]\n"
                            f"{summary}\n\n"
                            "저장 위치:\n"
                            f"{os.path.basename(out_path)}"
                        )
                        # Fix parent and wrap in try-except
                        try:
                            # Show custom preview if exists
                            if preview is not None:
                                show_preview_dialog(self, out_path, summary, preview)
                            else:
                                messagebox.showinfo("성공", msg, parent=self)
                        except:
                            # Fallback if parent invalid, though self should be valid
                            messagebox.showinfo("성공", msg)

                self.after(500, close_and_show_result)
                
            except BaseException as e:
                # Catch ALL exceptions including SystemExit/KeyboardInterrupt to log them
                result["error"] = str(e)
                _log_ui(f"Worker Error: {e}")
                
                # Log to Desktop for guaranteed visibility
                try:
                    log_path = os.path.join(os.path.expanduser("~"), "Desktop", "EasyMatch_Error_Log.txt")
                    with open(log_path, "a", encoding="utf-8") as f:
                        import datetime
                        f.write(f"\n[{datetime.datetime.now()}] CRITICAL RUNTIME ERROR: {str(e)}\n")
                        import traceback
                        traceback.print_exc(file=f)
                except: pass

                def show_error():
                    if progress_win.winfo_exists():
                        progress_win.destroy()
                    
                    msg = str(e)
                    if "xlwings" in msg.lower():
                        msg += (
                            "\n\n[힌트]\n"
                            "- Excel 설치 확인\n"
                            "- macOS: 개인정보 보호 및 보안 > 자동화에서 Terminal/iTerm2의 Excel 제어 허용\n"
                            "- 파일 모드로 사용 가능"
                        )
                    
                    # Force parent to main window (self) to ensure visibility
                    try:
                        messagebox.showerror("오류", f"실행 중 오류:\n{msg}", parent=self)
                    except:
                        messagebox.showerror("오류", f"실행 중 오류:\n{msg}")

                    self._log(f"Error: {msg}")
                
                self.after(0, show_error)
        
        # Start worker thread
        thread = threading.Thread(target=worker_thread, daemon=True)
        thread.start()
        
        # Handle window close
        def on_closing():
            cancel_flag["cancelled"] = True
            progress_win.destroy()
        
        progress_win.protocol("WM_DELETE_WINDOW", on_closing)